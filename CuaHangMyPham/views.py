from django.db import connection
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.core.paginator import Paginator
from django.http import HttpResponseRedirect
from django.http import Http404
from django.http import JsonResponse
from django.urls import reverse
from django.contrib import messages
from django.db.models import Sum, Count
from datetime import datetime
from openpyxl import Workbook
from django.contrib.auth.hashers import make_password
from django.contrib.auth.hashers import check_password
from django.core.files.storage import FileSystemStorage

# vnpay
import hashlib
import hmac
import json
import urllib.parse
import urllib.request
import random
import requests
from datetime import datetime
from django.conf import settings
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render, redirect
from urllib.parse import quote as urlquote
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import JsonResponse


from CuaHangMyPham.forms import PaymentForm
from CuaHangMyPham.vnpay import vnpay

# #---------------------------------------------------------------------------------------------------------------------#
# #--------------------------------------------------------TRANG CỬA HÀNG-----------------------------------------------#
# #---------------------------------------------------------------------------------------------------------------------#
def dangKyTaiKhoan(request):
    # Nếu người dùng đã đăng nhập, đăng xuất trước khi cho phép đăng ký tài khoản mới
    if request.session.get('tenTaiKhoan'):
        request.session.flush()

    error_message = None
    if request.method == 'POST':
        # Lấy dữ liệu từ form
        tenTaiKhoan = request.POST.get('TenTK')
        matKhau = request.POST.get('MatKhau')
        matKhau2 = request.POST.get('MatKhau2')
        email = request.POST.get('Email')
        vaiTro = 'user'  # Đặt vai trò mặc định là "user"

        # Kiểm tra mật khẩu nhập lại
        if matKhau != matKhau2:
            error_message = 'Mật khẩu và nhập lại mật khẩu không khớp'
        else:
            # Mã hóa mật khẩu
            hashed_password = make_password(matKhau)
            # Thực hiện lệnh INSERT
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO TaiKhoan (TenTK, MatKhau, Email, VaiTro)
                    VALUES (%s, %s, %s, %s)
                    """,
                    [tenTaiKhoan, hashed_password, email, vaiTro]
                )
                return redirect('dangNhap')  # Chuyển hướng đến trang đăng nhập sau khi đăng ký thành công

    return render(request, 'page/dangKyTaiKhoan.html', {'error_message': error_message})

def dangNhap(request):
    # Kiểm tra nếu người dùng đã đăng nhập
    if request.session.get('tenTaiKhoan'):
        return redirect('index')

    error_message = None
    if request.method == 'POST':
        # Lấy dữ liệu từ form
        tenTaiKhoan = request.POST.get('TenTK')
        matKhau = request.POST.get('MatKhau')

        # Kiểm tra tài khoản và mật khẩu
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT TenTK, MatKhau, VaiTro FROM TaiKhoan
                WHERE TenTK = %s
                """,
                [tenTaiKhoan]
            )
            taiKhoan = cursor.fetchone()

        if taiKhoan and check_password(matKhau, taiKhoan[1]):
            request.session['tenTaiKhoan'] = taiKhoan[0]
            request.session['vaiTro'] = taiKhoan[2]
            print(taiKhoan)
            print(taiKhoan[0])
            print(taiKhoan[2])
            # Nếu là admin thì chuyển hướng đến trang admin
            if taiKhoan[2] == 'admin':
                return redirect('trangChuAdmin')
            else:
                return redirect('index')
        else:
            error_message = 'Tên tài khoản hoặc mật khẩu không đúng'

    return render(request, 'page/dangNhap.html', {'error_message': error_message})

def dangXuat(request):
    request.session.flush()
    return redirect('dangNhap')

def is_logged_in(request):
    return {
        'tenTaiKhoan': request.session.get('tenTaiKhoan', False),
        'vaiTro': request.session.get('vaiTro', False),
    }


def session_info(request):
    session_info = {
        'tenTaiKhoan': request.session.get('tenTaiKhoan', 'Not set'),
        'vaiTro': request.session.get('vaiTro', 'Not set'),
    }
    return render(request, 'page/session_info.html', {'session_info': session_info})



def xemThongTinKhachHang(request):
    # Lấy thông tin khách hàng từ session
    tenTaiKhoan = request.session.get('tenTaiKhoan')
    if not tenTaiKhoan:
        return redirect('dangNhap')

    return render(request, 'page/customer_info.html')





# # #--------------------------------------------------HIỂN THỊ TRANG CHỦ-----------------------------------------------# 
def index(request):
    # Truy vấn 5 sản phẩm có ID lớn nhất (tương đương với top sản phẩm mới)
    with connection.cursor() as cursor:
        cursor.execute("""
           SELECT sp.MaSP, sp.TenSP, sp.HinhAnh, sp.GiaBan, COALESCE(AVG(dg.SoSao), 0) AS TrungBinhSoSao,
            ROUND(COALESCE(AVG(dg.SoSao), 0), 0) AS SoSao, COALESCE(km.PhanTramGiam, 0) AS PhanTramGiam,
            sp.GiaBan * (COALESCE(km.PhanTramGiam, 0) / 100.0) AS TienGiam, sp.GiaBan * (1 - (COALESCE(km.PhanTramGiam, 0) / 100.0)) AS GiaSauKM
            FROM SanPham sp LEFT JOIN DanhGia dg ON sp.MaSP = dg.MaSP
            LEFT JOIN CTKhuyenMai ct ON sp.MaSP = ct.MaSP
            LEFT JOIN KhuyenMai km ON ct.MaKM = km.MaKM AND CURRENT_DATE BETWEEN ct.NgayBD AND ct.NgayKT
            GROUP BY sp.MaSP, sp.TenSP, sp.HinhAnh, sp.GiaBan, km.PhanTramGiam
            ORDER BY sp.MaSP DESC
            LIMIT 5
        """)
        top_id_products = cursor.fetchall()

    # Phân trang cho sản phẩm mới
    items_per_page = 5  # Số sản phẩm trên mỗi trang
    paginator_top_id = Paginator(top_id_products, items_per_page)

    page_number_top_id = request.GET.get('page')
    page_obj_top_id = paginator_top_id.get_page(page_number_top_id)

    # Trả dữ liệu cho template
    data = {
        'topProNew': page_obj_top_id,
    }
    print(data)

    return render(request, 'page/index.html', data)


def xemTatCaSanPham(request):
    gia_khoang = request.GET.get('gia_khoang', None)  # Chỉ lấy một giá trị
    sap_xep = request.GET.get('sap_xep', None)  # Sắp xếp giá (tăng hoặc giảm)

    # Chuyển đổi khoảng giá thành điều kiện SQL
    gia_conditions = ""
    if gia_khoang == 'duoi_100':
        gia_conditions = "sp.GiaBan < 100000"
    elif gia_khoang == '100_200':
        gia_conditions = "sp.GiaBan BETWEEN 100000 AND 200000"
    elif gia_khoang == '200_300':
        gia_conditions = "sp.GiaBan BETWEEN 200000 AND 300000"
    elif gia_khoang == '300_500':
        gia_conditions = "sp.GiaBan BETWEEN 300000 AND 500000"
    elif gia_khoang == '500_1000':
        gia_conditions = "sp.GiaBan BETWEEN 500000 AND 1000000"
    elif gia_khoang == 'tren_1000':
        gia_conditions = "sp.GiaBan > 1000000"

    try:
        with connection.cursor() as cursor:
            query = """
                SELECT sp.MaSP, sp.TenSP, sp.HinhAnh, sp.GiaBan, COALESCE(AVG(dg.SoSao), 0) AS TrungBinhSoSao,
                ROUND(COALESCE(AVG(dg.SoSao), 0), 0) AS SoSao, COALESCE(km.PhanTramGiam, 0) AS PhanTramGiam,
                sp.GiaBan * (COALESCE(km.PhanTramGiam, 0) / 100.0) AS TienGiam, sp.GiaBan * (1 - (COALESCE(km.PhanTramGiam, 0) / 100.0)) AS GiaSauKM
                FROM SanPham sp 
                LEFT JOIN DanhGia dg ON sp.MaSP = dg.MaSP
                LEFT JOIN CTKhuyenMai ct ON sp.MaSP = ct.MaSP
                LEFT JOIN KhuyenMai km ON ct.MaKM = km.MaKM AND CURRENT_DATE BETWEEN ct.NgayBD AND ct.NgayKT
            """
            # Thêm điều kiện lọc giá
            if gia_conditions:
                query += " WHERE " + gia_conditions

            # Nhóm và sắp xếp
            query += """
                GROUP BY sp.MaSP, sp.TenSP, sp.HinhAnh, sp.GiaBan, km.PhanTramGiam
            """
            if sap_xep == 'tang_dan':
                query += " ORDER BY sp.GiaBan ASC"
            elif sap_xep == 'giam_dan':
                query += " ORDER BY sp.GiaBan DESC"

            cursor.execute(query)
            san_phams = cursor.fetchall()
    except Exception as e:
        print(f"Error when fetching product details: {e}")
        return HttpResponse("Có lỗi xảy ra khi xem sản phẩm.")

    return render(request, 'page/shop.html', {
        'san_phams': san_phams, 
        'gia_khoang': gia_khoang,  # Truyền giá trị khoảng giá đã chọn
        'sap_xep': sap_xep,  # Truyền thông tin sắp xếp
    })


def xemTatCaSanPhamTheoLoai(request, MaLoai):
    SanPham = None
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT sp.MaSP, sp.TenSP, sp.HinhAnh, sp.GiaBan, COALESCE(AVG(dg.SoSao), 0) AS TrungBinhSoSao,
                ROUND(COALESCE(AVG(dg.SoSao), 0), 0) AS SoSao, COALESCE(km.PhanTramGiam, 0) AS PhanTramGiam,
                sp.GiaBan * (COALESCE(km.PhanTramGiam, 0) / 100.0) AS TienGiam, sp.GiaBan * (1 - (COALESCE(km.PhanTramGiam, 0) / 100.0)) AS GiaSauKM
                FROM SanPham sp LEFT JOIN DanhGia dg ON sp.MaSP = dg.MaSP
                LEFT JOIN CTKhuyenMai ct ON sp.MaSP = ct.MaSP
                LEFT JOIN KhuyenMai km ON ct.MaKM = km.MaKM AND CURRENT_DATE BETWEEN ct.NgayBD AND ct.NgayKT
                GROUP BY sp.MaSP, sp.TenSP, sp.HinhAnh, sp.GiaBan, km.PhanTramGiam
                HAVING sp.MaLoai = %s
            """, [MaLoai])

            SanPham = cursor.fetchall()
            
    except Exception as e:
        print(f"Error when fetching product details: {e}")
        return HttpResponse("Có lỗi xảy ra khi xem sản phẩm.")

    if SanPham is None:
        return HttpResponse("Không tìm thấy sản phẩm.")

    return render(request, 'page/shop_theoloai.html', {
        'SanPhams': SanPham, 
    })



def xemTatCaSanPhamTheoThuongHieu(request, MaTH):
    SanPham = None
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT sp.MaSP, sp.TenSP, sp.HinhAnh, sp.GiaBan, COALESCE(AVG(dg.SoSao), 0) AS TrungBinhSoSao,
                ROUND(COALESCE(AVG(dg.SoSao), 0), 0) AS SoSao, COALESCE(km.PhanTramGiam, 0) AS PhanTramGiam,
                sp.GiaBan * (COALESCE(km.PhanTramGiam, 0) / 100.0) AS TienGiam, sp.GiaBan * (1 - (COALESCE(km.PhanTramGiam, 0) / 100.0)) AS GiaSauKM
                FROM SanPham sp LEFT JOIN DanhGia dg ON sp.MaSP = dg.MaSP
                LEFT JOIN CTKhuyenMai ct ON sp.MaSP = ct.MaSP
                LEFT JOIN KhuyenMai km ON ct.MaKM = km.MaKM AND CURRENT_DATE BETWEEN ct.NgayBD AND ct.NgayKT
                GROUP BY sp.MaSP, sp.TenSP, sp.HinhAnh, sp.GiaBan, km.PhanTramGiam
                HAVING sp.MaTH = %s
            """, [MaTH])

            SanPham = cursor.fetchall()
            
    except Exception as e:
        print(f"Error when fetching product details: {e}")
        return HttpResponse("Có lỗi xảy ra khi xem sản phẩm.")

    if SanPham is None:
        return HttpResponse("Không tìm thấy sản phẩm.")

    return render(request, 'page/shop_ThuongHieu.html', {
        'SanPhams': SanPham, 
    })



# # #------------------------------------------HIỂN THỊ CHI TIẾT SẢN PHẨM-----------------------------------------------# 
def xemChiTietCua1SanPham(request, MaSP):
    TenTk= request.session.get('tenTaiKhoan')
    SanPham = None
    DSDanhGia = []
    if request.method == 'POST':
        SoSao=request.POST.get('star')
    else:
        SoSao= None
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT sp.MaSP, sp.TenSP, HinhAnh, NSX, HSD, MoTa, GiaBan, SoLuongTon, TrongLuong, lsp.TenLoai, th.TenTH
                FROM SanPham sp
                JOIN LoaiSanPhan lsp ON sp.MaLoai = lsp.MaLoai
                JOIN ThuongHieu th ON sp.MaTH = th.MaTH
                WHERE sp.MaSP = %s
            """, [MaSP])
            SanPham = cursor.fetchone()
            if SoSao:
                DSDanhGia = xemDanhGiaCua1SanPham(MaSP, SoSao)
            else:
                DSDanhGia = xemDanhGiaCua1SanPham(MaSP)
            SoBaiDG= demSoBaiDanhGiaCua1SanPham(MaSP)
            SaoTB= xemSoSaoTrungBinhCua1SanPham(MaSP)
            if SaoTB is None:
                SaoTB = 0
            KtraDG= kiemTraDieuKienDanhGia(TenTk, MaSP)
            Km = xemThongTinKhuyenMai(MaSP)
            print(f"Khuyến mãi: {Km}")

    except Exception as e:
        print(f"Error when fetching product details: {e}")
        return HttpResponse("Có lỗi xảy ra khi xem chi tiết sản phẩm.")

    if SanPham is None:
        return HttpResponse("Không tìm thấy sản phẩm.")

    return render(request, 'page/shop_detail.html', {
        'SanPham': SanPham, 
        'DSDanhGia': DSDanhGia,
        'SoBaiDG': SoBaiDG,
        'SaoTB' : SaoTB,
        'KtraDG' : KtraDG,
        'KM' : Km
    })


def xemThongTinKhuyenMai(MaSP):
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT km.PhanTramGiam, sp.GiaBan *(km.PhanTramGiam / 100.0) AS TienGiam, (sp.GiaBan * (1 - (km.PhanTramGiam / 100.0))) AS GiaSauKM
                FROM SanPham sp
                JOIN CTKhuyenMai ct ON sp.MaSP = ct.MaSP
                JOIN KhuyenMai km ON ct.MaKM = km.MaKM
                WHERE sp.MaSP = %s 
                AND CURRENT_DATE BETWEEN ct.NgayBD AND ct.NgayKT;
            """, [MaSP])

            km = cursor.fetchone()
            if km:
                # Trả về dictionary với các tên trường thay vì tuple
                return {
                    'GiaGiam': km[0],  # Giảm giá (%)
                    'TienGiam': km[1],  # Tiền giảm
                    'GiaSauKM': km[2],  # Giá sau khuyến mãi
                }
            else:
                return None
    except Exception as e:
        print(f"Error when fetching discount info: {e}")
        return None
    

def xemDanhGiaCua1SanPham(MaSP, SoSao=None):
    try:
        with connection.cursor() as cursor:
            if SoSao:  # Nếu có giá trị SoSao (lọc theo sao)
                cursor.execute("""
                    SELECT hd.TenTK, dg.NoiDung, dg.NgayDanhGia, dg.SoSao
                    FROM DanhGia dg
                    JOIN HoaDon hd ON dg.MaHD = hd.MaHD
                    WHERE dg.MaSP = %s AND dg.SoSao = %s
                """, [MaSP, SoSao])
            else:  # Nếu không có giá trị SoSao (lấy tất cả đánh giá)
                cursor.execute("""
                    SELECT hd.TenTK, dg.NoiDung, dg.NgayDanhGia, dg.SoSao
                    FROM DanhGia dg
                    JOIN HoaDon hd ON dg.MaHD = hd.MaHD
                    WHERE dg.MaSP = %s
                """, [MaSP])

            # Lấy kết quả trả về
            DSDanhGia = cursor.fetchall()

    except Exception as e:
        print(f"Error when fetching reviews: {e}")
        return []  # Nếu có lỗi thì trả về danh sách rỗng

    return DSDanhGia  # Trả về danh sách đánh giá




# # #---------------------------------------------------ĐÁNH GIÁ SẢN PHẨM-----------------------------------------------#  

def xemDanhGiaCua1SanPham(MaSP, SoSao=None):
    try:
        with connection.cursor() as cursor:
            if SoSao:  # Nếu có giá trị SoSao (lọc theo sao)
                cursor.execute("""
                    SELECT hd.TenTK, dg.NoiDung, dg.NgayDanhGia, dg.SoSao
                    FROM DanhGia dg
                    JOIN HoaDon hd ON dg.MaHD = hd.MaHD
                    WHERE dg.MaSP = %s AND dg.SoSao = %s
                """, [MaSP, SoSao])
            else:  # Nếu không có giá trị SoSao (lấy tất cả đánh giá)
                cursor.execute("""
                    SELECT hd.TenTK, dg.NoiDung, dg.NgayDanhGia, dg.SoSao
                    FROM DanhGia dg
                    JOIN HoaDon hd ON dg.MaHD = hd.MaHD
                    WHERE dg.MaSP = %s
                """, [MaSP])

            # Lấy kết quả trả về
            DSDanhGia = cursor.fetchall()

    except Exception as e:
        print(f"Error when fetching reviews: {e}")
        return []  # Nếu có lỗi thì trả về danh sách rỗng

    return DSDanhGia  # Trả về danh sách đánh giá


def xemSoSaoTrungBinhCua1SanPham(MaSP):
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT AVG(SoSao)
                FROM DanhGia
                WHERE MaSP = %s
            """, [MaSP])

            saoTB = cursor.fetchone()[0] 
            if saoTB is not None: 
                saoTB = int(saoTB + 0.5)
            print(saoTB)
    except Exception as e:
        print(f"Error when fetching reviews: {e}")
        return []  
    return saoTB  


def demSoBaiDanhGiaCua1SanPham(MaSP):
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT COUNT(*) 
                FROM DanhGia 
                WHERE MaSP = %s
            """, [MaSP])

            soBaiDG = cursor.fetchone()[0]
    except Exception as e:
        print(f"Error when fetching reviews: {e}")
        return []  
    return soBaiDG  


def kiemTraDieuKienDanhGia(tenTK, MaSP):
    try:
        with connection.cursor() as cursor:
            # Lấy ra MaHD của người dùng nếu họ đã đặt hàng và chưa đánh giá sản phẩm đó trong vòng 1 tháng cho sản phẩm đó
            #-- Lấy ra MaHD mua sản phẩm đó của người dùng trong vòng 1 tháng
            #-- Kiểm tra xem MaSP và MaHD đã tồn tại trong bảng đánh giá chưa
            cursor.execute("""
                    SELECT hd.MaHD
                    FROM HoaDon hd
                    WHERE hd.TenTK = %s
                    AND hd.NgayLap >= DATE('now', '-1 month')
                    AND EXISTS (
                        SELECT 1 
                        FROM CTHoaDon cthd
                        WHERE cthd.MaHD = hd.MaHD AND cthd.MaSP = %s
                    )
                    AND NOT EXISTS (
                        SELECT 1
                        FROM DanhGia dg
                        WHERE dg.MaHD = hd.MaHD   
                        AND dg.MaSP = %s
                    )
            """, [tenTK, MaSP, MaSP])

            maHD = cursor.fetchone()  # Lấy MaHD nếu có
            if not maHD:
                return False
            return maHD
    except Exception as e:
        return HttpResponse(f"Lỗi khi kiểm tra đơn hàng: {e}")
    

    
def danhGiaMotSanPham(request, MaSP):
    if request.method == 'POST':
        noiDung = request.POST.get('NoiDung')
        ngayDG = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        soSao = request.POST.get('SoSao')
        tenTK = request.session.get('tenTaiKhoan')

        try:
            maHD= kiemTraDieuKienDanhGia(tenTK, MaSP)
            if maHD is False:
                return HttpResponse("Bạn không thể đánh giá sản phẩm này.")
        except Exception as e:
            return HttpResponse(f"Lỗi khi kiểm tra đơn hàng: {e}")

        # Nếu có MaHD, tiến hành thêm đánh giá vào bảng DanhGia
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    INSERT INTO DanhGia (MaSP, MaHD, NoiDung, NgayDanhGia, SoSao)
                    VALUES (%s, %s, %s, %s, %s)
                """, [MaSP, maHD[0], noiDung, ngayDG, soSao])  # maHD[0] là MaHD từ kết quả truy vấn
            return redirect(f'/shop_detail/{MaSP}')
        except Exception as e:
            return HttpResponse(f"Lỗi khi thêm đánh giá: {e}")
    return render(request, f'/shop_detail/danhGia/{MaSP}')


# # #------------------------------------------HIỂN THỊ THÔNG TIN NGƯỜI DÙNG-----------------------------------------------# 

def themThongTinKhachHang(request):
    tenTK = request.session.get('tenTaiKhoan')

    if request.method == 'POST':
        query = "SELECT MAX(CAST(SUBSTRING(MaKH, 3, LENGTH(MaKH) - 2) AS UNSIGNED)) FROM ThongTinKhachHang"
        maKH = taoMaTuDong(query)
        hoTen = request.POST.get('HoTen')
        sdt = request.POST.get('SDT')
        diaChi = request.POST.get('DiaChi')
        # macDinh = request.POST.get('MacDinh')

        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    INSERT INTO ThongTinKhachHang (MaKH, HoTen, SDT, DiaChi, TenTK)
                    VALUES (%s, %s, %s, %s, %s)
                """, [maKH, hoTen, sdt, diaChi, tenTK]) 
            return redirect(f'/xemThongTinKhachHang/')
        except Exception as e:
            return HttpResponse(f"Lỗi khi cập nhật thông tin khách hàng: {e}")
    return render(request, f'/xemThongTinKhachHang/')


def capNhatThongTinKhachHang(request):
    tenTK = request.session.get('tenTaiKhoan')

    if request.method == 'POST':
        maKH = request.POST.get('MaKH')
        hoTen = request.POST.get('HoTen')        
        sdt = request.POST.get('SDT')
        diaChi = request.POST.get('DiaChi')

        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    UPDATE ThongTinKhachHang Set HoTen= %s, SDT = %s, DiaChi= %s, TenTK= %s
                    WHERE MaKH= %s
                """, [hoTen, sdt, diaChi, tenTK, maKH])
            return redirect(f'/xemThongTinKhachHang/')
        except Exception as e:
            return HttpResponse(f"Lỗi khi cập nhật thông tin khách hàng: {e}")
    return render(request, f'/xemThongTinKhachHang/')



def xoaThongTinKhachHang(request):
    # tenTK = request.session.get('tenTaiKhoan')

    if request.method == 'POST':
        maKH = request.POST.get('MaKH')
    
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    DELETE FROM ThongTinKhachHang WHERE MaKH= %s
                """, [maKH])
            return redirect(f'/xemThongTinKhachHang/')
        except Exception as e:
            return HttpResponse(f"Lỗi khi cập nhật thông tin khách hàng: {e}")
    return render(request, f'/xemThongTinKhachHang/')


def xemThongTinCua1NguoiDung(request): 
    tenTK = request.session.get('tenTaiKhoan')
    
    if not tenTK:
        return render(request, 'page/thongTinKH.html', {
            'DSThongTinKH': [],
            'error': 'Bạn cần đăng nhập để xem thông tin.'
        })

    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT ttkh.HoTen, tk.Email, ttkh.SDT, ttkh.DiaChi, ttkh.MaKH, tk.TenTK
                FROM TaiKhoan tk
                JOIN ThongTinKhachHang ttkh ON tk.TenTK = ttkh.TenTK
                WHERE tk.TenTK = %s
            """, [tenTK])
  
            DSThongTinKH = cursor.fetchall()

            if not DSThongTinKH:
                return render(request, 'page/thongTinKH.html', {
                    'DSThongTinKH': [],
                    'error': 'Không tìm thấy thông tin người dùng.'
                })

    except Exception as e:
        return render(request, 'page/thongTinKH.html', {
            'DSThongTinKH': [],
            'error': 'Đã xảy ra lỗi trong quá trình lấy thông tin.'
        })

    return render(request, 'page/thongTinKH.html', {
        'DSThongTinKH': DSThongTinKH
    })


def kiemTraDieuKienCapNhat(email, username, passCu):
    try:
        with connection.cursor() as cursor:
            cursor.execute(""" SELECT MatKhau FROM TaiKhoan 
                WHERE TenTK = %s and Email = %s                    
            """, [username, email])

            thongTin = cursor.fetchone()
            if not thongTin:
                return False
            
            if thongTin and check_password(passCu, thongTin[0]):
                return True
            return False
    except Exception as e:
        return HttpResponse(f"Lỗi khi kiểm tra đơn hàng: {e}")
    

def capNhatTaiKhoan(request):
    if request.method == 'POST':
        email = request.POST.get('Email')
        username = request.POST.get('Username')
        pass_cu = request.POST.get('Passcu')
        pass_moi = request.POST.get('Passmoi')

        try:
            with connection.cursor() as cursor:
                # Kiểm tra mật khẩu cũ
                if not kiemTraDieuKienCapNhat(email, username, pass_cu):
                    return JsonResponse({'success': False, 'message': 'Thông tin tài khoản hoặc mật khẩu cũ không đúng.'})
                
                # Mã hóa mật khẩu mới
                hashed_password = make_password(pass_moi)
                
                # Cập nhật tài khoản
                cursor.execute("""
                    UPDATE TaiKhoan 
                    SET Email = %s, MatKhau = %s
                    WHERE TenTK = %s
                """, [email, hashed_password, username])
                
                return JsonResponse({'success': True, 'message': 'Cập nhật tài khoản thành công.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Lỗi khi cập nhật tài khoản: {e}'})

    return JsonResponse({'success': False, 'message': 'Phương thức không hợp lệ.'})


def xemlichSuMuaHang(request): 
    tenTK = request.session.get('tenTaiKhoan')
    KtraDG=[]
    KtraDG1=[]
    DSCTHoaDon1 = []
    DSCTHoaDon = []
    if not tenTK:
        return render(request, 'page/thongTinKH.html', {
            'DSHoaDon': [],
            'error': 'Bạn cần đăng nhập để xem thông tin.'
        })

    try:
        #đơn hàng chưa xử lý
        with connection.cursor() as cursor:
            cursor.execute("""
               SELECT hd.MaHD, hd.NgayLap, hd.TongTien, hd.TrangThai, hd.PhuongThucThanhToan ,  hd.TrangThaiThanhToan,
                hd.TenKhachHang, hd.SDT, hd.DiaChiGiaoHang
                FROM HoaDon hd
                WHERE hd.TenTK = %s and hd.TrangThai=0
            """, [tenTK])
            DSHoaDon = cursor.fetchall()

            if DSHoaDon:
                cursor.execute("""
                SELECT hd.MaHD, cd.MaSP , sp.TenSP , sp.HinhAnh, sp.GiaBan, cd.DonGia, cd.SoLuong , hd.TongTien
                FROM HoaDon hd , CTHoaDon cd, SanPham sp 
                WHERE hd.MaHD = cd.MaHD  
                AND cd.MaSP = sp.MaSP 
                AND hd.TenTK = %s
                """, [tenTK])
                DSCTHoaDon = cursor.fetchall()    

                #lặp DSCTHoaDon, đưa MaSP trong DSCTHoaDon vào hàm kiemTraDieuKienDanhGia,
                #nếu hàm trả về true thì lưu list , list này gồm 2 phần tử, 1 là MaHD, MaSP, và thuộc tính True/False
                 # Lặp qua danh sách chi tiết hóa đơn
                for chiTiet in DSCTHoaDon:
                    maHD = chiTiet[0]
                    maSP = chiTiet[1]

                    # Gọi hàm kiểm tra điều kiện đánh giá
                    isEligible = kiemTraDieuKienDanhGia(tenTK, maSP)
                    if isEligible:
                        dg= True
                    else:
                        dg= False
                    # Lưu kết quả vào danh sách
                    print("XEM NÈ: ", maHD, maSP, dg)
                    KtraDG.append({
                        'MaHD': maHD,
                        'MaSP': maSP,
                        'Eligible': dg
                    }) 

        #Đơn hàng đã xử lý
        with connection.cursor() as cursor:
            cursor.execute("""
               SELECT hd.MaHD, hd.NgayLap, hd.TongTien, hd.TrangThai, hd.PhuongThucThanhToan ,  hd.TrangThaiThanhToan,
                hd.TenKhachHang, hd.SDT, hd.DiaChiGiaoHang
                FROM HoaDon hd
                WHERE hd.TenTK = %s and hd.TrangThai=1
            """, [tenTK])
            DSHoaDon1 = cursor.fetchall()

            if DSHoaDon1:
                cursor.execute("""
                SELECT hd.MaHD, cd.MaSP , sp.TenSP , sp.HinhAnh, sp.GiaBan, cd.DonGia, cd.SoLuong , hd.TongTien
                FROM HoaDon hd , CTHoaDon cd , CTThanhToan ct , SanPham sp 
                WHERE hd.MaHD = cd.MaHD 
                AND cd.MaHD = ct.MaHD  
                AND cd.MaSP = sp.MaSP 
                AND hd.TenTK = %s
                """, [tenTK])
                DSCTHoaDon1 = cursor.fetchall()    

                #lặp DSCTHoaDon, đưa MaSP trong DSCTHoaDon vào hàm kiemTraDieuKienDanhGia,
                #nếu hàm trả về true thì lưu list , list này gồm 2 phần tử, 1 là MaHD, MaSP, và thuộc tính True/False
                 # Lặp qua danh sách chi tiết hóa đơn
                for chiTiet in DSCTHoaDon1:
                    maHD = chiTiet[0]
                    maSP = chiTiet[1]

                    # Gọi hàm kiểm tra điều kiện đánh giá
                    isEligible = kiemTraDieuKienDanhGia(tenTK, maSP)
                    if isEligible:
                        dg= True
                    else:
                        dg= False
                    # Lưu kết quả vào danh sách
                    print("XEM NÈ: ", maHD, maSP, dg)
                    KtraDG1.append({
                        'MaHD': maHD,
                        'MaSP': maSP,
                        'Eligible': dg
                    }) 

    except Exception as e:
        return render(request, 'page/lichSuMuaHang.html', {
            'DSHoaDon':[],
            'DSCTHoaDon': [],
            'KtraDG':[],
            'DSHoaDon1':[],
            'DSCTHoaDon1': [],
            'KtraDG1':[],
            'error': 'Đã xảy ra lỗi trong quá trình lấy thông tin.'
        })

    return render(request, 'page/lichSuMuaHang.html', {
        'DSHoaDon': DSHoaDon,
        'DSCTHoaDon': DSCTHoaDon,
        'KtraDG' :KtraDG,
        'DSHoaDon1': DSHoaDon1,
        'DSCTHoaDon1': DSCTHoaDon1,
        'KtraDG1' :KtraDG1

    })


# # #---------------------------------------------------THANH TOÁN ĐƠN HÀNG-----------------------------------------------#
def taoMaHD(query):
    cursor = connection.cursor()
    cursor.execute(query)
    result = cursor.fetchone()
    cursor.close()

    if result[0]:
        max_number = int(result[0]) + 1
    else:
        max_number = 1  # Nếu chưa có khuyến mãi nào, bắt đầu từ 1

    # Tạo mã mới với tiền tố "KM" và số có 3 chữ số (ví dụ: "KM001")
    return f"HD{str(max_number).zfill(3)}"

# def checkout(request):
#     # user = request.session.get('user')  # Lấy người dùng hiện tại từ session
#     # user = 'TK001'
#     user = request.session.get('tenTaiKhoan')

#     # Lấy giỏ hàng từ cơ sở dữ liệu
#     with connection.cursor() as cursor:
#         cursor.execute("""
#             SELECT c.MaSP, p.TenSP, p.GiaBan, c.SoLuong
#             FROM CTGioHang c
#             INNER JOIN SanPham p ON c.MaSP = p.MaSP
#             WHERE c.TenTK = %s
#         """, [user])
#         cart_items = cursor.fetchall()

#     cart_details = []
#     total_amount = 0
#     for index, item in enumerate(cart_items, start=1):
#         product_id, product_name, product_price, quantity = item[0], item[1], item[2], item[3]
#         product_total = product_price * quantity
#         total_amount += product_total

#         cart_details.append({
#             'index': index,
#             'product_name': product_name,
#             'quantity': quantity,
#             'product_total': "{:,.0f}".format(product_total).replace(",", ".")
#         })

#         # Lấy danh sách địa chỉ nhận hàng từ cơ sở dữ liệu
#     with connection.cursor() as cursor:
#         cursor.execute("""
#             SELECT HoTen, DiaChi, SDT
#             FROM ThongTinKhachHang
#             WHERE TenTK = %s
#         """, [user])
#         delivery_addresses = cursor.fetchall()
    

#     if request.method == 'POST':

#         use_existing_address = request.POST.get('use_existing_address')
#         if use_existing_address:
#             existing_address = request.POST.get('existing_address')
#             if not existing_address:
#                 return render(request, 'page/checkout.html', {
#                     'cart_details': cart_details,
#                     'total_amount': total_amount,
#                     'delivery_addresses': delivery_addresses,
#                     'error': 'Bạn cần chọn một địa chỉ nhận hàng từ danh sách.'
#                 })
#             recipient_name, recipient_address, recipient_phone = existing_address.split(', ')
#             ward = district = city = ''  # Assuming these are not needed for existing addresses
#         else:
#             recipient_name = request.POST.get('recipient_name')
#             ward = request.POST.get('ward')
#             district = request.POST.get('district')
#             city = request.POST.get('city')
#             recipient_address = request.POST.get('recipient_address')
#             recipient_phone = request.POST.get('recipient_phone')

#         query = "SELECT MAX(CAST(SUBSTRING(MaHD, 3, LENGTH(MaHD) - 2) AS UNSIGNED)) FROM HoaDon"
#         mahd = taoMaHD(query)
#         recipient_name = request.POST.get('recipient_name')
#         ward = request.POST.get('ward')
#         district = request.POST.get('district')
#         city = request.POST.get('city')
#         recipient_address = request.POST.get('recipient_address')
#         recipient_phone = request.POST.get('recipient_phone')
#         method_payment = request.POST.get('method_payment')

#         trangthaithanhtoan = "Chưa thanh toán"
#         trangthai = "0"
#         ngayLap = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#         full_recipient_address = f"{recipient_address}, {ward}, {district}, {city}"

#         # Tạo đơn hàng mới
#         with connection.cursor() as cursor:
#             cursor.execute("""
#                 INSERT INTO HoaDon (MaHD, TenTK, TenKhachHang, DiaChiGiaoHang, SDT, PhuongThucThanhToan, TongTien, TrangThaiThanhToan, TrangThai, NgayLap)
#                 VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING MaHD
#             """, [mahd , user, recipient_name, full_recipient_address, recipient_phone, method_payment, 0, trangthaithanhtoan, trangthai, ngayLap])
#             order_id = cursor.fetchone()[0]

#             # Thêm chi tiết đơn hàng và tính tổng tiền
#             total_amount = 0
#             for item in cart_items:
#                 # product_id, product_price, quantity = item
#                 total_amount += item[3] * item[2]

#                 cursor.execute("""
#                     INSERT INTO CTHoaDon (MaHD, MaSP, SoLuong, DonGia)
#                     VALUES (%s, %s, %s, %s)
#                 """, [order_id, item[0], item[3], item[2]])

#             # Cập nhật tổng số tiền đơn hàng
#             cursor.execute("""
#                 UPDATE HoaDon
#                 SET TongTien = %s
#                 WHERE MaHD = %s
#             """, [total_amount, order_id])

        
#         if method_payment == "Thanh toán VNPay":
#             ipaddr = get_client_ip(request)
#             vnp = vnpay()
#             vnp.requestData['vnp_Version'] = '2.1.0'
#             vnp.requestData['vnp_Command'] = 'pay'
#             vnp.requestData['vnp_TmnCode'] = settings.VNPAY_TMN_CODE
#             vnp.requestData['vnp_Amount'] = total_amount * 100
#             vnp.requestData['vnp_CurrCode'] = 'VND'
#             vnp.requestData['vnp_TxnRef'] = mahd
#             vnp.requestData['vnp_OrderInfo'] = "Thanh toán hóa đơn"
#             vnp.requestData['vnp_OrderType'] = "billpayment"
#             # Check language, default: vn
#             # if language and language != '':
#             #     vnp.requestData['vnp_Locale'] = language
#             # else:
#             vnp.requestData['vnp_Locale'] = 'vn'
#                 # Check bank_code, if bank_code is empty, customer will be selected bank on VNPAY
#             # if bank_code and bank_code != "":
#             # vnp.requestData['vnp_BankCode'] = "NCB"

#             vnp.requestData['vnp_CreateDate'] = datetime.now().strftime('%Y%m%d%H%M%S')  # 20150410063022
#             vnp.requestData['vnp_IpAddr'] = ipaddr
#             vnp.requestData['vnp_ReturnUrl'] = f"{settings.VNPAY_RETURN_URL}?user={user}"
#             vnpay_payment_url = vnp.get_payment_url(settings.VNPAY_PAYMENT_URL, settings.VNPAY_HASH_SECRET_KEY)
#             # print(vnpay_payment_url)

#             return redirect(vnpay_payment_url)
#         else: #thanh toán COD
#             with connection.cursor() as cursor:

#                 # trừ sl tồn trong kho
#                 for item in cart_items:
#                     product_id = item[0]
#                     quantity = item[3]
#                     cursor.execute("""
#                         UPDATE SanPham SET SoLuongTon = SoLuongTon - %s WHERE MaSP = %s
#                     """, [quantity, product_id])
                
#                 # Xóa giỏ hàng sau khi đặt hàng thành công
#                 cursor.execute("""
#                     DELETE FROM CTGioHang WHERE TenTK = %s
#                 """, [user])
            


#         return HttpResponseRedirect('/order_details')

#     total_amount = "{:,.0f}".format(total_amount).replace(",", ".")  # Format tổng tiền

#     return render(request, 'page/checkout.html', {
#         'cart_details': cart_details,
#         'total_amount': total_amount,
#         'delivery_addresses': delivery_addresses
#     })

def checkout(request):
    user = request.session.get('tenTaiKhoan')

    # Lấy giỏ hàng từ cơ sở dữ liệu
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT c.MaSP, p.TenSP, p.GiaBan, c.SoLuong
            FROM CTGioHang c
            INNER JOIN SanPham p ON c.MaSP = p.MaSP
            WHERE c.TenTK = %s
        """, [user])
        cart_items = cursor.fetchall()

    cart_details = []
    total_amount = 0
    for index, item in enumerate(cart_items, start=1):
        product_id, product_name, product_price, quantity = item[0], item[1], item[2], item[3]
        product_total = product_price * quantity
        total_amount += product_total

        cart_details.append({
            'index': index,
            'product_name': product_name,
            'quantity': quantity,
            'product_total': "{:,.0f}".format(product_total).replace(",", ".")
        })

    # Lấy danh sách địa chỉ nhận hàng từ cơ sở dữ liệu
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT HoTen, DiaChi, SDT
            FROM ThongTinKhachHang
            WHERE TenTK = %s
        """, [user])
        delivery_addresses = cursor.fetchall()

    if request.method == 'POST':
        use_existing_address = request.POST.get('use_existing_address')
        if use_existing_address:
            existing_address = request.POST.get('existing_address')
            if not existing_address:
                return render(request, 'page/checkout.html', {
                    'cart_details': cart_details,
                    'total_amount': total_amount,
                    'delivery_addresses': delivery_addresses,
                    'error': 'Bạn cần chọn một địa chỉ nhận hàng từ danh sách.'
                })
            recipient_name, recipient_address, recipient_phone = existing_address.split(', ')
            ward = district = city = ''  # Assuming these are not needed for existing addresses
            full_recipient_address = recipient_address  # Use only the recipient_address for existing addresses
        else:
            recipient_name = request.POST.get('recipient_name')
            ward = request.POST.get('ward')
            district = request.POST.get('district')
            city = request.POST.get('city')
            recipient_address = request.POST.get('recipient_address')
            recipient_phone = request.POST.get('recipient_phone')
            full_recipient_address = f"{recipient_address}, {ward}, {district}, {city}"

        query = "SELECT MAX(CAST(SUBSTRING(MaHD, 3, LENGTH(MaHD) - 2) AS UNSIGNED)) FROM HoaDon"
        mahd = taoMaHD(query)
        method_payment = request.POST.get('method_payment')

        trangthaithanhtoan = "Chưa thanh toán"
        trangthai = "0"
        ngayLap = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Tạo đơn hàng mới
        with connection.cursor() as cursor:
            cursor.execute("""
                INSERT INTO HoaDon (MaHD, TenTK, TenKhachHang, DiaChiGiaoHang, SDT, PhuongThucThanhToan, TongTien, TrangThaiThanhToan, TrangThai, NgayLap)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING MaHD
            """, [mahd, user, recipient_name, full_recipient_address, recipient_phone, method_payment, 0, trangthaithanhtoan, trangthai, ngayLap])
            order_id = cursor.fetchone()[0]

            # Thêm chi tiết đơn hàng và tính tổng tiền
            total_amount = 0
            for item in cart_items:
                total_amount += item[3] * item[2]

                cursor.execute("""
                    INSERT INTO CTHoaDon (MaHD, MaSP, SoLuong, DonGia)
                    VALUES (%s, %s, %s, %s)
                """, [order_id, item[0], item[3], item[2]])

            # Cập nhật tổng số tiền đơn hàng
            cursor.execute("""
                UPDATE HoaDon
                SET TongTien = %s
                WHERE MaHD = %s
            """, [total_amount, order_id])

        if method_payment == "Thanh toán VNPay":
            ipaddr = get_client_ip(request)
            vnp = vnpay()
            vnp.requestData['vnp_Version'] = '2.1.0'
            vnp.requestData['vnp_Command'] = 'pay'
            vnp.requestData['vnp_TmnCode'] = settings.VNPAY_TMN_CODE
            vnp.requestData['vnp_Amount'] = total_amount * 100
            vnp.requestData['vnp_CurrCode'] = 'VND'
            vnp.requestData['vnp_TxnRef'] = mahd
            vnp.requestData['vnp_OrderInfo'] = "Thanh toán hóa đơn"
            vnp.requestData['vnp_OrderType'] = "billpayment"
            vnp.requestData['vnp_Locale'] = 'vn'
            vnp.requestData['vnp_CreateDate'] = datetime.now().strftime('%Y%m%d%H%M%S')
            vnp.requestData['vnp_IpAddr'] = ipaddr
            vnp.requestData['vnp_ReturnUrl'] = f"{settings.VNPAY_RETURN_URL}?user={user}"
            vnpay_payment_url = vnp.get_payment_url(settings.VNPAY_PAYMENT_URL, settings.VNPAY_HASH_SECRET_KEY)

            return redirect(vnpay_payment_url)
        else:  # thanh toán COD
            with connection.cursor() as cursor:
                # trừ sl tồn trong kho
                for item in cart_items:
                    product_id = item[0]
                    quantity = item[3]
                    cursor.execute("""
                        UPDATE SanPham SET SoLuongTon = SoLuongTon - %s WHERE MaSP = %s
                    """, [quantity, product_id])

                # Xóa giỏ hàng sau khi đặt hàng thành công
                cursor.execute("""
                    DELETE FROM CTGioHang WHERE TenTK = %s
                """, [user])

        return HttpResponseRedirect('/lichSuMuaHang')

    total_amount = "{:,.0f}".format(total_amount).replace(",", ".")  # Format tổng tiền

    return render(request, 'page/checkout.html', {
        'cart_details': cart_details,
        'total_amount': total_amount,
        'delivery_addresses': delivery_addresses
    })

def payment_return(request):
    user = request.GET.get('user')

    inputData = request.GET
    if inputData:
        vnp = vnpay()
        vnp.responseData = inputData.dict()
        order_id = inputData['vnp_TxnRef']
        amount = int(inputData['vnp_Amount']) / 100
        order_desc = inputData['vnp_OrderInfo']
        vnp_TransactionNo = inputData['vnp_TransactionNo']
        vnp_ResponseCode = inputData['vnp_ResponseCode']
        vnp_TmnCode = inputData['vnp_TmnCode']
        vnp_PayDate = inputData['vnp_PayDate']
        vnp_BankCode = inputData['vnp_BankCode']
        vnp_CardType = inputData['vnp_CardType']

        # Chuỗi thời gian từ vnp_PayDate
        vnp_PayDate = inputData['vnp_PayDate']

        # Chuyển đổi thành đối tượng datetime
        pay_date = datetime.strptime(vnp_PayDate, '%Y%m%d%H%M%S')

        # Định dạng lại để hiển thị (ví dụ: 'YYYY-MM-DD HH:MM:SS')
        formatted_pay_date = pay_date.strftime('%Y-%m-%d %H:%M:%S')

        # thêm vào chi tiết thanh toán

        if vnp.validate_response(settings.VNPAY_HASH_SECRET_KEY):
            if vnp_ResponseCode == "00":

                with connection.cursor() as cursor:
                    
                    cursor.execute("""
                        INSERT INTO CTThanhToan (MaHD, MaGD, TrangThaiGiaoDich, CongThanhToan, NgayThanhToan, GhiChu)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, [order_id , vnp_TransactionNo, "Thành công", "VNPay", formatted_pay_date, order_desc])

                    cursor.execute("""
                        SELECT c.MaSP, p.TenSP, p.GiaBan, c.SoLuong
                        FROM CTGioHang c
                        INNER JOIN SanPham p ON c.MaSP = p.MaSP
                        WHERE c.TenTK = %s
                    """, [user])
                    cart_items = cursor.fetchall()

                    # trừ sl tồn trong kho
                    for item in cart_items:
                        product_id = item[0]
                        quantity = item[3]
                        cursor.execute("""
                            UPDATE SanPham SET SoLuongTon = SoLuongTon - %s WHERE MaSP = %s
                        """, [quantity, product_id])
                    
                    # Xóa giỏ hàng sau khi đặt hàng thành công
                    cursor.execute("""
                        DELETE FROM CTGioHang WHERE TenTK = %s
                    """, [user])

                    # update trang thai thanh toan
                    cursor.execute("""
                            UPDATE HoaDon SET TrangThaiThanhToan = "Đã thanh toán" WHERE MaHD = %s
                        """, [order_id])

                return render(request, "page/payment/payment_return.html", {"title": "Kết quả thanh toán",
                                                            "result": "Thành công", "order_id": order_id,
                                                            "amount": amount,
                                                            "order_desc": order_desc,
                                                            "vnp_TransactionNo": vnp_TransactionNo,
                                                            "vnp_ResponseCode": vnp_ResponseCode}) # trả về trang thành công
            else:
                return render(request, "page/payment/payment_return.html", {"title": "Kết quả thanh toán",
                                                            "result": "Lỗi", "order_id": order_id,
                                                            "amount": amount,
                                                            "order_desc": order_desc,
                                                            "vnp_TransactionNo": vnp_TransactionNo,
                                                            "vnp_ResponseCode": vnp_ResponseCode})
        else:
            return render(request, "page/payment/payment_return.html",
                        {"title": "Kết quả thanh toán", "result": "Lỗi", "order_id": order_id, "amount": amount,
                        "order_desc": order_desc, "vnp_TransactionNo": vnp_TransactionNo,
                        "vnp_ResponseCode": vnp_ResponseCode, "msg": "Sai checksum"})
    else:
        return render(request, "page/payment/payment_return.html", {"title": "Kết quả thanh toán", "result": ""})


    # Hóa đơn thanh toán



def payment(request):

    if request.method == 'POST':
        # Process input data and build url payment
        form = PaymentForm(request.POST)
        if form.is_valid():
            order_type = form.cleaned_data['order_type']
            order_id = form.cleaned_data['order_id']
            amount = form.cleaned_data['amount']
            order_desc = form.cleaned_data['order_desc']
            bank_code = form.cleaned_data['bank_code']
            language = form.cleaned_data['language']
            ipaddr = get_client_ip(request)
            # Build URL Payment
            vnp = vnpay()
            vnp.requestData['vnp_Version'] = '2.1.0'
            vnp.requestData['vnp_Command'] = 'pay'
            vnp.requestData['vnp_TmnCode'] = settings.VNPAY_TMN_CODE
            vnp.requestData['vnp_Amount'] = amount * 100
            vnp.requestData['vnp_CurrCode'] = 'VND'
            vnp.requestData['vnp_TxnRef'] = order_id
            vnp.requestData['vnp_OrderInfo'] = order_desc
            vnp.requestData['vnp_OrderType'] = order_type
            # Check language, default: vn
            if language and language != '':
                vnp.requestData['vnp_Locale'] = language
            else:
                vnp.requestData['vnp_Locale'] = 'vn'
                # Check bank_code, if bank_code is empty, customer will be selected bank on VNPAY
            if bank_code and bank_code != "":
                vnp.requestData['vnp_BankCode'] = bank_code

            vnp.requestData['vnp_CreateDate'] = datetime.now().strftime('%Y%m%d%H%M%S')  # 20150410063022
            vnp.requestData['vnp_IpAddr'] = ipaddr
            vnp.requestData['vnp_ReturnUrl'] = settings.VNPAY_RETURN_URL
            vnpay_payment_url = vnp.get_payment_url(settings.VNPAY_PAYMENT_URL, settings.VNPAY_HASH_SECRET_KEY)
            print(vnpay_payment_url)
            return redirect(vnpay_payment_url)
        else:
            print("Form input not validate")
    else:
        return render(request, "page/payment/payment.html", {"title": "Thanh toán"})


def index_payment(request):
    return render(request, "page/payment/index.html", {"title": "Danh sách demo"})


def hmacsha512(key, data):
    byteKey = key.encode('utf-8')
    byteData = data.encode('utf-8')
    return hmac.new(byteKey, byteData, hashlib.sha512).hexdigest()


def payment_ipn(request):
    inputData = request.GET
    if inputData:
        vnp = vnpay()
        vnp.responseData = inputData.dict()
        order_id = inputData['vnp_TxnRef']
        amount = inputData['vnp_Amount']
        order_desc = inputData['vnp_OrderInfo']
        vnp_TransactionNo = inputData['vnp_TransactionNo']
        vnp_ResponseCode = inputData['vnp_ResponseCode']
        vnp_TmnCode = inputData['vnp_TmnCode']
        vnp_PayDate = inputData['vnp_PayDate']
        vnp_BankCode = inputData['vnp_BankCode']
        vnp_CardType = inputData['vnp_CardType']
        if vnp.validate_response(settings.VNPAY_HASH_SECRET_KEY):
            # Check & Update Order Status in your Database
            # Your code here
            firstTimeUpdate = True
            totalamount = True
            if totalamount:
                if firstTimeUpdate:
                    if vnp_ResponseCode == '00':
                        print('Payment Success. Your code implement here')
                    else:
                        print('Payment Error. Your code implement here')

                    # Return VNPAY: Merchant update success
                    result = JsonResponse({'RspCode': '00', 'Message': 'Confirm Success'})
                else:
                    # Already Update
                    result = JsonResponse({'RspCode': '02', 'Message': 'Order Already Update'})
            else:
                # invalid amount
                result = JsonResponse({'RspCode': '04', 'Message': 'invalid amount'})
        else:
            # Invalid Signature
            result = JsonResponse({'RspCode': '97', 'Message': 'Invalid Signature'})
    else:
        result = JsonResponse({'RspCode': '99', 'Message': 'Invalid request'})

    return result


# def payment_return(request):
#     inputData = request.GET
#     if inputData:
#         vnp = vnpay()
#         vnp.responseData = inputData.dict()
#         order_id = inputData['vnp_TxnRef']
#         amount = int(inputData['vnp_Amount']) / 100
#         order_desc = inputData['vnp_OrderInfo']
#         vnp_TransactionNo = inputData['vnp_TransactionNo']
#         vnp_ResponseCode = inputData['vnp_ResponseCode']
#         vnp_TmnCode = inputData['vnp_TmnCode']
#         vnp_PayDate = inputData['vnp_PayDate']
#         vnp_BankCode = inputData['vnp_BankCode']
#         vnp_CardType = inputData['vnp_CardType']
#         if vnp.validate_response(settings.VNPAY_HASH_SECRET_KEY):
#             if vnp_ResponseCode == "00":
#                 return render(request, "page/payment/payment_return.html", {"title": "Kết quả thanh toán",
#                                                             "result": "Thành công", "order_id": order_id,
#                                                             "amount": amount,
#                                                             "order_desc": order_desc,
#                                                             "vnp_TransactionNo": vnp_TransactionNo,
#                                                             "vnp_ResponseCode": vnp_ResponseCode})
#             else:
#                 return render(request, "page/payment/payment_return.html", {"title": "Kết quả thanh toán",
#                                                             "result": "Lỗi", "order_id": order_id,
#                                                             "amount": amount,
#                                                             "order_desc": order_desc,
#                                                             "vnp_TransactionNo": vnp_TransactionNo,
#                                                             "vnp_ResponseCode": vnp_ResponseCode})
#         else:
#             return render(request, "page/payment/payment_return.html",
#                         {"title": "Kết quả thanh toán", "result": "Lỗi", "order_id": order_id, "amount": amount,
#                         "order_desc": order_desc, "vnp_TransactionNo": vnp_TransactionNo,
#                         "vnp_ResponseCode": vnp_ResponseCode, "msg": "Sai checksum"})
#     else:
#         return render(request, "page/payment/payment_return.html", {"title": "Kết quả thanh toán", "result": ""})


def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

n = random.randint(10**11, 10**12 - 1)
n_str = str(n)
while len(n_str) < 12:
    n_str = '0' + n_str


def query(request):
    if request.method == 'GET':
        return render(request, "page/payment/query.html", {"title": "Kiểm tra kết quả giao dịch"})

    url = settings.VNPAY_API_URL
    secret_key = settings.VNPAY_HASH_SECRET_KEY
    vnp_TmnCode = settings.VNPAY_TMN_CODE
    vnp_Version = '2.1.0'

    vnp_RequestId = n_str
    vnp_Command = 'querydr'
    vnp_TxnRef = request.POST['order_id']
    vnp_OrderInfo = 'kiem tra gd'
    vnp_TransactionDate = request.POST['trans_date']
    vnp_CreateDate = datetime.now().strftime('%Y%m%d%H%M%S')
    vnp_IpAddr = get_client_ip(request)

    hash_data = "|".join([
        vnp_RequestId, vnp_Version, vnp_Command, vnp_TmnCode,
        vnp_TxnRef, vnp_TransactionDate, vnp_CreateDate,
        vnp_IpAddr, vnp_OrderInfo
    ])

    secure_hash = hmac.new(secret_key.encode(), hash_data.encode(), hashlib.sha512).hexdigest()

    data = {
        "vnp_RequestId": vnp_RequestId,
        "vnp_TmnCode": vnp_TmnCode,
        "vnp_Command": vnp_Command,
        "vnp_TxnRef": vnp_TxnRef,
        "vnp_OrderInfo": vnp_OrderInfo,
        "vnp_TransactionDate": vnp_TransactionDate,
        "vnp_CreateDate": vnp_CreateDate,
        "vnp_IpAddr": vnp_IpAddr,
        "vnp_Version": vnp_Version,
        "vnp_SecureHash": secure_hash
    }

    headers = {"Content-Type": "application/json"}

    response = requests.post(url, headers=headers, data=json.dumps(data))

    if response.status_code == 200:
        response_json = json.loads(response.text)
    else:
        response_json = {"error": f"Request failed with status code: {response.status_code}"}

    return render(request, "page/payment/query.html", {"title": "Kiểm tra kết quả giao dịch", "response_json": response_json})

def refund(request):
    if request.method == 'GET':
        return render(request, "page/payment/refund.html", {"title": "Hoàn tiền giao dịch"})

    url = settings.VNPAY_API_URL
    secret_key = settings.VNPAY_HASH_SECRET_KEY
    vnp_TmnCode = settings.VNPAY_TMN_CODE
    vnp_RequestId = n_str
    vnp_Version = '2.1.0'
    vnp_Command = 'refund'
    vnp_TransactionType = request.POST['TransactionType']
    vnp_TxnRef = request.POST['order_id']
    vnp_Amount = request.POST['amount']
    vnp_OrderInfo = request.POST['order_desc']
    vnp_TransactionNo = '0'
    vnp_TransactionDate = request.POST['trans_date']
    vnp_CreateDate = datetime.now().strftime('%Y%m%d%H%M%S')
    vnp_CreateBy = 'user01'
    vnp_IpAddr = get_client_ip(request)

    hash_data = "|".join([
        vnp_RequestId, vnp_Version, vnp_Command, vnp_TmnCode, vnp_TransactionType, vnp_TxnRef,
        vnp_Amount, vnp_TransactionNo, vnp_TransactionDate, vnp_CreateBy, vnp_CreateDate,
        vnp_IpAddr, vnp_OrderInfo
    ])

    secure_hash = hmac.new(secret_key.encode(), hash_data.encode(), hashlib.sha512).hexdigest()

    data = {
        "vnp_RequestId": vnp_RequestId,
        "vnp_TmnCode": vnp_TmnCode,
        "vnp_Command": vnp_Command,
        "vnp_TxnRef": vnp_TxnRef,
        "vnp_Amount": vnp_Amount,
        "vnp_OrderInfo": vnp_OrderInfo,
        "vnp_TransactionDate": vnp_TransactionDate,
        "vnp_CreateDate": vnp_CreateDate,
        "vnp_IpAddr": vnp_IpAddr,
        "vnp_TransactionType": vnp_TransactionType,
        "vnp_TransactionNo": vnp_TransactionNo,
        "vnp_CreateBy": vnp_CreateBy,
        "vnp_Version": vnp_Version,
        "vnp_SecureHash": secure_hash
    }

    headers = {"Content-Type": "application/json"}

    response = requests.post(url, headers=headers, data=json.dumps(data))

    if response.status_code == 200:
        response_json = json.loads(response.text)
    else:
        response_json = {"error": f"Request failed with status code: {response.status_code}"}

    return render(request, "page/payment/refund.html", {"title": "Kết quả hoàn tiền giao dịch", "response_json": response_json})



####-------------------------------GIỎ HÀNG-----------------------------
# # #hiển thị giỏ hàng của 1 người
def cart(request):
    tenTaiKhoan = request.session.get('tenTaiKhoan')
    if tenTaiKhoan:
        # Câu truy vấn SQL để lấy sản phẩm và số lượng trong giỏ hàng
        query = """
        SELECT sp.MaSP, sp.TenSP, sp.HinhAnh, sp.GiaBan, sp.SoLuongTon, 
               IFNULL(ct.SoLuong, 0) AS SoLuongTrongGioHang
        FROM SanPham sp
        LEFT JOIN CTGioHang ct ON sp.MaSP = ct.MaSP
        WHERE ct.TenTK = %s
        ORDER BY sp.MaSP DESC;
        """

        # Sử dụng connection.cursor() để thực thi câu truy vấn SQL
        with connection.cursor() as cursor:
            cursor.execute(query, [tenTaiKhoan])
            # Lấy tất cả kết quả trả về dưới dạng tuple
            cart_items = cursor.fetchall()

        # Chuyển đổi kết quả thành danh sách cart_details
        cart_details = []
        total_price = 0  # Khởi tạo biến tổng tiền

        for item in cart_items:
            item_dict = {
                'MaSP': item[0],
                'TenSP': item[1],
                'HinhAnh': item[2],
                'GiaBan': item[3],
                'SoLuongTon': item[4],
                'SoLuongTrongGioHang': item[5],
                'ItemTotal': item[3] * item[5]  # Tính tổng tiền cho từng sản phẩm
            }
            cart_details.append(item_dict)
            total_price += item_dict['ItemTotal']  # Cộng tổng tiền cho tất cả sản phẩm

        # Lấy số lượng sản phẩm trong giỏ hàng
        cart_item_count = get_cart_item_count(tenTaiKhoan)

        return render(request, 'page/cart.html', {'cart_details': cart_details, 'total_price': total_price, 'cart_item_count': cart_item_count})
    else:
        # Nếu không có thông tin người dùng trong session, trả về trang đăng nhập
        return redirect('dangNhap')
# #thêm sản phẩm vô giỏ hàng
def addProToCart(request):
    # Gán trực tiếp user_id (ví dụ, user_id = "tk001")
    tenTaiKhoan = request.session.get('tenTaiKhoan')
    if not tenTaiKhoan:
        return redirect('dangNhap')

    # Kiểm tra action và id_product trong GET request
    if request.GET.get('action') == 'addcart' and 'id_product' in request.GET:
        product_id = request.GET['id_product']
        
        # Kiểm tra xem user_id có giá trị không
        if tenTaiKhoan:  # Nếu user_id có giá trị (ở đây là "tk001")
            # Câu truy vấn SQL để kiểm tra xem sản phẩm đã có trong giỏ hàng chưa
            query = """
            SELECT * FROM CTGioHang WHERE TenTK = %s AND MaSP = %s;
            """
            with connection.cursor() as cursor:
                cursor.execute(query, [tenTaiKhoan, product_id])
                cart_item = cursor.fetchone()  # Lấy kết quả trả về (nếu có)

            if cart_item:  # Nếu sản phẩm đã có trong giỏ hàng, cập nhật số lượng
                # Cập nhật số lượng giỏ hàng
                update_query = """
                UPDATE CTGioHang
                SET SoLuong = SoLuong + 1
                WHERE TenTK = %s AND MaSP = %s;
                """
                with connection.cursor() as cursor:
                    cursor.execute(update_query, [tenTaiKhoan, product_id])
            else:
                # Nếu sản phẩm chưa có trong giỏ hàng, thêm mới
                insert_query = """
                INSERT INTO CTGioHang (TenTK, MaSP, SoLuong, DonGia)
                SELECT %s, %s, 1, GiaBan FROM SanPham WHERE MaSP = %s;
                """
                with connection.cursor() as cursor:
                    cursor.execute(insert_query, [tenTaiKhoan, product_id, product_id])

            # Sau khi thao tác xong, chuyển hướng về trang giỏ hàng
            return redirect('cart')  # Redirect đến trang giỏ hàng

    #     else:
    #         # Nếu không có thông tin người dùng trong session, chuyển hướng tới trang đăng nhập
    #         return redirect('login')  # Redirect đến trang đăng nhập

    # # Nếu không có action 'addcart' hoặc 'id_product', chuyển hướng về trang cửa hàng
    # return redirect('shop')  # Redirect về trang cửa hàng
# #xóa 1 sản phẩm trong giỏ hàng
def deleteProFromCart(request, product_id):
    tenTaiKhoan = request.session.get('tenTaiKhoan')
    if not tenTaiKhoan:
        return redirect('dangNhap')

    if tenTaiKhoan:
        # Câu truy vấn SQL để xóa sản phẩm khỏi giỏ hàng
        delete_query = """
        DELETE FROM CTGioHang 
        WHERE TenTK = %s AND MaSP = %s;
        """
        with connection.cursor() as cursor:
            cursor.execute(delete_query, [tenTaiKhoan, product_id])

    return redirect('cart')  # Redirect về trang giỏ hàng
# #xóa tất cả sản phẩm giỏ hàng
def clearCart(request):
    # Lấy user_id từ session hoặc gán trực tiếp nếu cần
    tenTaiKhoan = request.session.get('tenTaiKhoan')

    if tenTaiKhoan:
        # Sử dụng TenTK thay vì user_id nếu đó là tên đúng cột trong bảng
        delete_query = """
        DELETE FROM CTGioHang
        WHERE TenTK = %s;
        """
        with connection.cursor() as cursor:
            cursor.execute(delete_query, [tenTaiKhoan])

    # Sau khi xóa xong, chuyển hướng về trang giỏ hàng
    return redirect('cart')
# #tăng số lượng sản phẩm trong giỏ hàng
def increase_quantity(request, product_id):
    # Gán user_id (hoặc lấy từ session nếu cần)
    tenTaiKhoan = request.session.get('tenTaiKhoan')

    if tenTaiKhoan:
        # Câu truy vấn SQL để tăng số lượng sản phẩm trong giỏ hàng
        update_query = """
        UPDATE CTGioHang
        SET SoLuong = SoLuong + 1
        WHERE TenTK = %s AND MaSP = %s;
        """
        with connection.cursor() as cursor:
            cursor.execute(update_query, [tenTaiKhoan, product_id])

    # Sau khi cập nhật số lượng, chuyển hướng về trang giỏ hàng
    return redirect('cart')
# #giảm số lượng sản phẩm trong giỏ hàng
def decrease_quantity(request, product_id):
    # Gán user_id (hoặc lấy từ session nếu cần)
    tenTaiKhoan = request.session.get('tenTaiKhoan')
    if tenTaiKhoan:
        # Kiểm tra số lượng hiện tại của sản phẩm trong giỏ hàng
        select_query = """
        SELECT SoLuong
        FROM CTGioHang
        WHERE TenTK = %s AND MaSP = %s;
        """
        with connection.cursor() as cursor:
            cursor.execute(select_query, [tenTaiKhoan, product_id])
            current_quantity = cursor.fetchone()

        if current_quantity and current_quantity[0] > 1:
            # Câu truy vấn SQL để giảm số lượng sản phẩm trong giỏ hàng
            update_query = """
            UPDATE CTGioHang
            SET SoLuong = SoLuong - 1
            WHERE TenTK = %s AND MaSP = %s;
            """
            with connection.cursor() as cursor:
                cursor.execute(update_query, [tenTaiKhoan, product_id])

    # Sau khi cập nhật số lượng, chuyển hướng về trang giỏ hàng
    return redirect('cart')
# #đếm số lượng sản phẩm trong giỏ hàng
def get_cart_item_count(user_id):
    # Gán trực tiếp user_id
    # user_id = "tk001"  # Hoặc lấy từ session nếu cần
    query = """
    SELECT COUNT(DISTINCT MaSP) AS total_items
    FROM CTGioHang
    WHERE TenTK = %s;
    """
    with connection.cursor() as cursor:
        cursor.execute(query, [user_id])
        result = cursor.fetchone()

    return result[0] if result and result[0] is not None else 0

#--------------------------------------TRANG ADMIN-----------------------------------#
#Trang chủ admin
def trangChuAdmin(request):
    # Kiểm tra xem người dùng đã đăng nhập chưa và có phải là admin không
    if not request.session.get('tenTaiKhoan'):
        return redirect('dangNhap')
    
    if request.session.get('vaiTro') != 'admin':
        return render(request, 'page/tuChoiTruyCap.html', {'message': 'Bạn không có quyền truy cập vào trang này.'})

    return render(request, 'admin/trangChuAdmin.html')



# #--------------------------------------------------------QUẢN LÝ NHHÂN VIÊN-----------------------------------------------# 

def xemDanhSachNhanVien(request):
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT nv.MaNV, nv.TenNV, nv.DiaChi, nv.Email, nv.SDT, nv.ChucVu, tk.TenTK, tk.MatKhau, tk.Email, tk.VaiTro
            FROM NhanVien nv
            LEFT JOIN TaiKhoan tk ON nv.TenTK = tk.TenTK
        """)
        listNhanVien = cursor.fetchall()

    # PHÂN TRANG
    items_per_page = 5  # Số nhân viên trên mỗi trang
    paginator = Paginator(listNhanVien, items_per_page)

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    total_pages = paginator.num_pages

    previous_page = max(page_obj.number - 1, 1)  # Trang trước đó
    next_page = min(page_obj.number + 1, total_pages)  # Trang tiếp theo

    start_page = max(1, min(page_obj.number - 1, total_pages - 2))  # Vị trí của nút trang đầu tiên
    end_page = min(total_pages, start_page + 2)

    page_range = range(start_page, end_page + 1)  # Create a range or list of page numbers

    data = {
        'listNhanVien': page_obj,
        'page_range': page_range,  # Pass the page range to the template
        'total_pages': total_pages,
        'previous_page': previous_page,
        'next_page': next_page,
    }

    return render(request, 'admin/danhSachNhanVien.html', data)

def capTaiKhoan(request, maNhanVien):
    if request.method == 'POST':
        tenTaiKhoan = request.POST.get('TenTK')
        matKhau = request.POST.get('MatKhau')
        email = request.POST.get('Email')
        vaiTro = request.POST.get('VaiTro')

        try:
            with connection.cursor() as cursor:
                # Kiểm tra xem tài khoản đã tồn tại hay chưa
                cursor.execute("SELECT COUNT(*) FROM TaiKhoan WHERE TenTK = %s", [tenTaiKhoan])
                exists = cursor.fetchone()[0]

                if exists:
                    # Nếu tài khoản đã tồn tại, cập nhật thông tin
                    cursor.execute(
                        """
                        UPDATE TaiKhoan
                        SET MatKhau = %s, Email = %s, VaiTro = %s
                        WHERE TenTK = %s
                        """,
                        [matKhau, email, vaiTro, tenTaiKhoan]
                    )
                else:
                    # Nếu tài khoản chưa tồn tại, thêm mới
                    cursor.execute(
                        """
                        INSERT INTO TaiKhoan (TenTK, MatKhau, Email, VaiTro)
                        VALUES (%s, %s, %s, %s)
                        """,
                        [tenTaiKhoan, matKhau, email, vaiTro]
                    )

                # Cập nhật thông tin tài khoản trong bảng NhanVien
                cursor.execute(
                    """
                    UPDATE NhanVien
                    SET TenTK = %s
                    WHERE MaNV = %s
                    """,
                    [tenTaiKhoan, maNhanVien]
                )
            return HttpResponseRedirect('/admin/danhSachNhanVien')
        except Exception as e:
            return HttpResponse('Đã có lỗi xảy ra: ' + str(e))
    else:
        #Thông báo lỗi
        return HttpResponse('Không thể cấp tài khoản cho nhân viên')


def taoMaNhanVien():
    cursor = connection.cursor()
    query = "SELECT MAX(CAST(SUBSTRING(MaNV, 3, LENGTH(MaNV) - 2) AS UNSIGNED)) FROM NhanVien"
    cursor.execute(query)
    result = cursor.fetchone()
    cursor.close()

    if result[0] is not None:
        max_id = int(result[0]) + 1
    else:
        max_id = 1  # Nếu không có dữ liệu, bắt đầu từ 1

    # Tạo mã nhân viên mới với định dạng NVxxx
    return f"NV{str(max_id).zfill(3)}"

def themNhanVien(request):
    if request.method == "POST":
        maNhanVien = taoMaNhanVien()
        tenNhanVien = request.POST.get('TenNV')
        diaChi = request.POST.get('DiaChi')
        sDT = request.POST.get('SDT')
        email = request.POST.get('Email')
        chucVu = request.POST.get('ChucVu')
        #tenTK = request.POST.get('TenTK')
    
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO NhanVien (MaNV, TenNV, DiaChi, SDT, Email, ChucVu)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                    [maNhanVien, tenNhanVien, diaChi, sDT, email, chucVu]
                )
            return HttpResponseRedirect('/admin/danhSachNhanVien')
        except Exception as e:
            return HttpResponse('Đã có lỗi xảy ra: ' + str(e))
    return render(request, 'admin/themNhanVien.html')

def suaNhanVien(request, maNhanVien):
    if request.method == 'POST':
        tenNhanVien = request.POST.get('TenNV')
        diaChi = request.POST.get('DiaChi')
        sDT = request.POST.get('SDT')
        email = request.POST.get('Email')
        chucVu = request.POST.get('ChucVu')

        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    UPDATE NhanVien
                    SET TenNV = %s, DiaChi = %s, SDT = %s, Email = %s, ChucVu = %s
                    WHERE MaNV = %s
                    """,
                    [tenNhanVien, diaChi, sDT, email, chucVu, maNhanVien]
                )
            return HttpResponseRedirect('/admin/danhSachNhanVien')
        except Exception as e:
            return HttpResponse('Đã có lỗi xảy ra: ' + str(e))
    else:
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM NhanVien WHERE MaNV = %s", [maNhanVien])
            nhanVien = cursor.fetchone()
            if nhanVien is None:
                return HttpResponse('Không tìm thấy nhân viên')

        data = {
            'MaNV': nhanVien[0],
            'TenNV': nhanVien[1],
            'DiaChi': nhanVien[2],
            'SDT': nhanVien[3],
            'Email': nhanVien[4],
            'ChucVu': nhanVien[5],
        }
        return render(request, 'admin/suaNhanVien.html', data)
    
def xoaNhanVien(request, maNhanVien):
    if request.method == 'POST':
        try:
            with connection.cursor() as cursor:
                # Lấy TenTK của nhân viên
                cursor.execute("SELECT TenTK FROM NhanVien WHERE MaNV = %s", [maNhanVien])
                tenTaiKhoan = cursor.fetchone()
                
                # Xóa nhân viên
                cursor.execute("DELETE FROM NhanVien WHERE MaNV = %s", [maNhanVien])

                if tenTaiKhoan:
                    tenTaiKhoan = tenTaiKhoan[0]
                    # Xóa tài khoản của nhân viên
                    cursor.execute("DELETE FROM TaiKhoan WHERE TenTK = %s", [tenTaiKhoan])

            return HttpResponseRedirect('/admin/danhSachNhanVien')
        except Exception as e:
            return HttpResponse('Đã có lỗi xảy ra: ' + str(e))
    else:
        return HttpResponse('Không thể xóa nhân viên')

def xemChiTietNhanVien(request, maNhanVien):
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT nv.MaNV, nv.TenNV, nv.DiaChi, nv.Email, nv.SDT, nv.ChucVu, tk.TenTK, tk.MatKhau, tk.Email, tk.VaiTro
            FROM NhanVien nv
            LEFT JOIN TaiKhoan tk ON nv.TenTK = tk.TenTK
            WHERE nv.MaNV = %s
        """, [maNhanVien])
        nhanVien = cursor.fetchone()

    if nhanVien is None:
        return HttpResponse('Không tìm thấy nhân viên')

    data = {
        'MaNV': nhanVien[0],
        'TenNV': nhanVien[1],
        'DiaChi': nhanVien[2],
        'Email': nhanVien[3],
        'SDT': nhanVien[4],
        'ChucVu': nhanVien[5],
        'TenTK': nhanVien[6],
        'MatKhau': nhanVien[7],
        'EmailTK': nhanVien[8],
        'VaiTro': nhanVien[9],
    }
    return render(request, 'admin/xemThongTinNhanVien.html', data)

# #--------------------------------------------------------QUẢN LÝ TÀI KHOẢN-----------------------------------------------# 
#danh sách người dùng
def xemDanhSachTaiKhoan(request):
    # session_info = is_logged_in(request)
    # if not session_info['email'] or session_info['role'] != 'admin':
    #     return HttpResponseRedirect('/login')
    
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM TaiKhoan")
        listUser = cursor.fetchall()

    # PHÂN TRANG
    items_per_page = 5  # Số sản phẩm trên mỗi trang
    paginator = Paginator(listUser, items_per_page)

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    total_pages = paginator.num_pages

    previous_page = max(page_obj.number - 1, 1)  # Trang trước đó
    next_page = min(page_obj.number + 1, total_pages)  # Trang tiếp theo

    start_page = max(1, min(page_obj.number - 1, total_pages - 2))  # Vị trí của nút trang đầu tiên
    end_page = min(total_pages, start_page + 2)

    page_range = range(start_page, end_page + 1)  # Create a range or list of page numbers

    data = {
        'listUser': page_obj,
        'page_range': page_range,  # Pass the page range to the template
        'total_pages': total_pages,
        'previous_page': previous_page,
        'next_page': next_page,
    }
    return render(request, 'admin/danhSachTaiKhoan.html', data)

# #thêm tài khoản
def themTaiKhoan(request):
    if request.method == 'POST':
        # Lấy dữ liệu từ form
        tenTaiKhoan = request.POST.get('TenTK')
        matKhau = request.POST.get('MatKhau')
        email = request.POST.get('Email')
        vaiTro = request.POST.get('VaiTro')

        # Thực hiện lệnh INSERT
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO TaiKhoan (TenTK, MatKhau, Email, VaiTro)
                    VALUES (%s, %s, %s, %s)
                    """,
                    [tenTaiKhoan, matKhau, email, vaiTro]
                )
            return HttpResponseRedirect('/admin/danhSachTaiKhoan')
        except Exception as e:
            return HttpResponse('Đã có lỗi xảy ra: ' + str(e))
    return render(request, 'admin/themTaiKhoan.html')

def suaTaiKhoan(request, tenTaiKhoan):
    if request.method == 'POST':
        # Lấy dữ liệu từ form
        matKhau = request.POST.get('MatKhau')
        email = request.POST.get('Email')
        vaiTro = request.POST.get('VaiTro')

        # Thực hiện lệnh UPDATE
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    UPDATE TaiKhoan
                    SET MatKhau = %s, Email = %s, VaiTro = %s
                    WHERE TenTK = %s
                    """,
                    [matKhau, email, vaiTro, tenTaiKhoan]
                )
            return HttpResponseRedirect('/admin/danhSachTaiKhoan')
        except Exception as e:
            return HttpResponse('Đã có lỗi xảy ra: ' + str(e))
    else:
        # Lấy thông tin tài khoản hiện tại
        with connection.cursor() as cursor:
            cursor.execute("SELECT TenTK, MatKhau, Email, VaiTro FROM TaiKhoan WHERE TenTK = %s", [tenTaiKhoan])
            user = cursor.fetchone()
            if user is None:
                return HttpResponse('Không tìm thấy tài khoản')

        data = {
            'TenTK': user[0],
            'MatKhau': user[1],
            'Email': user[2],
            'VaiTro': user[3],
        }
        return render(request, 'admin/suaTaiKhoan.html', data)
    
def xoaTaiKhoan(request, tenTaiKhoan):
    if request.method == 'POST':
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM TaiKhoan WHERE TenTK = %s", [tenTaiKhoan])
        return HttpResponseRedirect('/admin/danhSachTaiKhoan')
    else:
        return HttpResponse('Không thể xóa tài khoản')

#-----------------------------------KHUYẾN MÃI----------------------------------
def taoMaTuDong(query):
    cursor = connection.cursor()
    cursor.execute(query)
    result = cursor.fetchone()
    cursor.close()

    if result[0]:
        max_number = int(result[0]) + 1
    else:
        max_number = 1  # Nếu chưa có khuyến mãi nào, bắt đầu từ 1

    # Tạo mã mới với tiền tố "KM" và số có 3 chữ số (ví dụ: "KM001")
    return f"KM{str(max_number).zfill(3)}"

def khuyenMai(request):
        #     session_info = is_logged_in(request)
#     if not session_info['email'] or session_info['role'] != 'admin':
#         return HttpResponseRedirect('/login')
    danh_sach_khuyen_mai = []

    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT MaKM, TenKM, PhanTramGiam FROM KhuyenMai")
            danh_sach_khuyen_mai = cursor.fetchall()  # Lấy toàn bộ kết quả
            
            # PHÂN TRANG
            items_per_page = 5  # Số sản phẩm trên mỗi trang
            paginator = Paginator(danh_sach_khuyen_mai, items_per_page)

            page_number = request.GET.get('page', 1)
            page_obj = paginator.get_page(page_number)

            total_pages = paginator.num_pages

            previous_page = max(page_obj.number - 1, 1)  # Trang trước đó
            next_page = min(page_obj.number + 1, total_pages)  # Trang tiếp theo

            start_page = max(1, min(page_obj.number - 1, total_pages - 2))  # Vị trí của nút trang đầu tiên
            end_page = min(total_pages, start_page + 2)

            page_range = range(start_page, end_page + 1)  # Create a range or list of page numbers

            data = {
                'listKM': page_obj,
                'page_range': page_range,  # Pass the page range to the template
                'total_pages': total_pages,
                'previous_page': previous_page,
                'next_page': next_page,
            }

    except Exception as e:
        return HttpResponse(f"Lỗi khi truy vấn khuyến mãi: {e}")
    
    
    # with connection.cursor() as cursor:
    #     cursor.execute("SELECT * FROM TaiKhoan")
    #     listUser = cursor.fetchall()

    # Truyền dữ liệu khuyến mãi đến template
    return render(request, 'admin/khuyenMai.html',data)


def themKhuyenMai(request):
    #     session_info = is_logged_in(request)
#     if not session_info['email'] or session_info['role'] != 'admin':
#         return HttpResponseRedirect('/login')
    
    if request.method == 'POST':
        # Lấy dữ liệu từ form
        query = "SELECT MAX(CAST(SUBSTRING(MaKM, 3, LENGTH(MaKM) - 2) AS UNSIGNED)) FROM KhuyenMai"
        ma_km = taoMaTuDong(query)
        ten_km = request.POST.get('TenKM')
        gia_giam = request.POST.get('PhanTramGiam')

        # Thực hiện lệnh INSERT
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO KhuyenMai (MaKM, TenKM, PhanTramGiam)
                    VALUES (%s, %s, %s)
                    """,
                    [ma_km, ten_km, gia_giam]
                )
            return redirect('/admin/khuyenMai')  # Điều hướng sau khi thêm thành công
        except Exception as e:
            return HttpResponse(f"Lỗi khi thêm khuyến mãi: {e}")
    return render(request, 'admin/themKhuyenMai.html')


def suaKhuyenMai(request, makm):
    if request.method == 'POST':
        ten_khuyen_mai = request.POST.get('TenKM')
        gia_giam = request.POST.get('PhanTramGiam')

        cursor = connection.cursor()
        query = """
            UPDATE KhuyenMai
            SET TenKM = %s, PhanTramGiam = %s
            WHERE MaKM = %s
        """
        cursor.execute(query, [ten_khuyen_mai, gia_giam, makm])
        cursor.close()

        # Chuyển hướng sau khi cập nhật thành công
        return HttpResponseRedirect('/admin/khuyenMai')

    # Lấy thông tin khuyến mãi hiện tại để hiển thị trong form
    cursor = connection.cursor()
    query = "SELECT MaKM, TenKM, PhanTramGiam FROM KhuyenMai WHERE MaKM = %s"
    cursor.execute(query, [makm])
    khuyen_mai = cursor.fetchone()
    cursor.close()

    # Truyền dữ liệu khuyến mãi vào template
    return render(request, 'admin/suaKhuyenMai.html', {'KhuyenMai': khuyen_mai})

def xoaKhuyenMai(request, makm):
    cursor = connection.cursor()
    query = "DELETE FROM KhuyenMai WHERE MaKM = %s"
    cursor.execute(query, [makm])
    cursor.close()

    # Chuyển hướng sau khi xóa thành công
    return HttpResponseRedirect('/admin/khuyenMai')


def chiTietKhuyenMai(request, makm):
    # Truy vấn thông tin chi tiết khuyến mãi
    sql_khuyen_mai = """SELECT km.MaKM, km.TenKM, km.PhanTramGiam, ct.NgayBD, ct.NgayKT
                        FROM KhuyenMai km
                        LEFT JOIN CTKhuyenMai ct ON km.MaKM = ct.MaKM
                        WHERE km.MaKM = %s"""
    sql_san_pham = """
        SELECT SanPham.MaSP, SanPham.TenSP, SanPham.GiaBan, CTKhuyenMai.NgayBD, CTKhuyenMai.NgayKT
        FROM SanPham
        JOIN CTKhuyenMai ON SanPham.MaSP = CTKhuyenMai.MaSP
        WHERE CTKhuyenMai.MaKM = %s
    """

    with connection.cursor() as cursor:
        # Lấy thông tin khuyến mãi
        cursor.execute(sql_khuyen_mai, [makm])
        row = cursor.fetchone()

        # Kiểm tra nếu không tìm thấy khuyến mãi
        if not row:
            raise Http404("Khuyến mãi không tồn tại.")

        # Lấy thông tin sản phẩm áp dụng khuyến mãi
        cursor.execute(sql_san_pham, [makm])
        products = cursor.fetchall()

    # Tạo dictionary để truyền vào template
    khuyen_mai = {
        'MaKM': row[0],
        'TenKM': row[1],
        'PhanTramGiam': row[2],
    }

    # Lọc thông tin sản phẩm
    san_pham_list = [
        {
            'MaSP': product[0],
            'TenSP': product[1],
            'GiaBan': product[2],
            'NgayBatDau': product[3],
            'NgayKetThuc': product[4],
        }
        for product in products
    ]

    # PHÂN TRANG
    items_per_page = 5  # Số sản phẩm trên mỗi trang
    paginator = Paginator(san_pham_list, items_per_page)

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    total_pages = paginator.num_pages

    previous_page = max(page_obj.number - 1, 1)  # Trang trước đó
    next_page = min(page_obj.number + 1, total_pages)  # Trang tiếp theo

    start_page = max(1, min(page_obj.number - 1, total_pages - 2))  # Vị trí của nút trang đầu tiên
    end_page = min(total_pages, start_page + 2)

    page_range = range(start_page, end_page + 1)  # Tạo dải số trang

    # Render thông tin ra template
    return render(request, 'admin/CTKhuyenMai.html', {
        'khuyen_mai': khuyen_mai,
        'lstKM': page_obj,
        'page_range': page_range,
        'total_pages': total_pages,
        'previous_page': previous_page,
        'next_page': next_page,
    })




#  CT khuyến mãi
def themSPVaoKhuyenMai(request, makm):
    # Lấy thông tin khuyến mãi từ mã
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM khuyenmai WHERE MaKM = %s", [makm])
    khuyen_mai = cursor.fetchone()

    if not khuyen_mai:
        raise Http404("Khuyến mãi không tồn tại.")
    
    # Lấy danh sách sản phẩm
    cursor.execute("""SELECT sp.MaSP, sp.TenSP, sp.GiaBan, ct.NgayBD, ct.NgayKT FROM SanPham sp
                    LEFT JOIN CTKhuyenMai ct ON sp.MaSP = ct.MaSP""")
    san_pham_list = cursor.fetchall()

    if request.method == 'POST':
        # Lấy danh sách sản phẩm đã chọn từ form
        selected_products = request.POST.getlist('selected_products')

        # Lấy ngày bắt đầu và ngày kết thúc từ form
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        if not start_date or not end_date:
            # Nếu ngày bắt đầu hoặc kết thúc không hợp lệ, trả về lỗi
            return render(request, 'admin/themCTKM.html', {
                'khuyen_mai': khuyen_mai,
                'san_pham_list': san_pham_list,
                'error_message': "Vui lòng nhập đầy đủ ngày bắt đầu và ngày kết thúc."
            })

        # Kiểm tra và thêm các sản phẩm đã chọn vào CTKhuyenMai
        for product_id in selected_products:
            # Kiểm tra xem sản phẩm đã tồn tại trong CTKhuyenMai chưa
            cursor.execute("""SELECT 1 FROM ctkhuyenmai WHERE MaKM = %s AND MaSP = %s""", [makm, product_id])
            if not cursor.fetchone():  # Nếu không tìm thấy thì thêm
                cursor.execute("""
                    INSERT INTO ctkhuyenmai (MaKM, MaSP, NgayBD, NgayKT)
                    VALUES (%s, %s, %s, %s)
                """, [makm, product_id, start_date, end_date])
            else:
                cursor.execute("""UPDATE CTKhuyenMai SET NgayBD = %s, NgayKT = %s
                            WHERE MaKM = %s AND MaSP = %s""", [start_date, end_date ,makm, product_id]) 

        return HttpResponseRedirect(reverse('CTKhuyenMai', kwargs={'makm': makm}))
    
    return render(request, 'admin/themCTKM.html', {'khuyen_mai': khuyen_mai, 'san_pham_list': san_pham_list})



def get_available_products(request, start_date, end_date):
    # Truy vấn sản phẩm chưa tham gia khuyến mãi trong khoảng thời gian
    cursor = connection.cursor()
    cursor.execute("""
        SELECT sp.MaSP, sp.TenSP, sp.GiaBan, ct.NgayBD, ct.NgayKT FROM SanPham sp LEFT JOIN CTKhuyenMai ct ON sp.MaSP = ct.MaSP
        WHERE sp.MaSP NOT IN (
            SELECT MaSP FROM CTKhuyenMai WHERE MaKM IN (
                SELECT MaKM FROM khuyenmai WHERE NOT (NgayKT < %s OR NgayBD > %s)
            )
        )
    """, [start_date, end_date])

    products = cursor.fetchall()
    # Kiểm tra xem có dữ liệu không
    if not products:
        print("No products found")
    else:
        print("Products:", products)

    product_data = [{'MaSP': product[0], 'TenSP': product[1], 'GiaSP': product[2], 'NgayBD': product[3], 'NgayKT': product[4]} for product in products]
    return JsonResponse({'products': product_data})


def xoaCTKhuyenMai(request, makm, masp):
    # Kiểm tra sự tồn tại của khuyến mãi và sản phẩm
    cursor = connection.cursor()
    
    # Kiểm tra xem mã khuyến mãi có tồn tại không
    cursor.execute("SELECT * FROM khuyenmai WHERE MaKM = %s", [makm])
    khuyen_mai = cursor.fetchone()
    if not khuyen_mai:
        raise Http404("Khuyến mãi không tồn tại.")
    
    # Kiểm tra xem sản phẩm có tồn tại trong CTKhuyenMai không
    cursor.execute("SELECT * FROM ctkhuyenmai WHERE MaKM = %s AND MaSP = %s", [makm, masp])
    ct_khuyen_mai = cursor.fetchone()
    if not ct_khuyen_mai:
        raise Http404("Sản phẩm không tồn tại trong chương trình khuyến mãi.")
    
    # Thực hiện xóa mục trong CTKhuyenMai
    cursor.execute("DELETE FROM ctkhuyenmai WHERE MaKM = %s AND MaSP = %s", [makm, masp])

    # Sau khi xóa, chuyển hướng lại trang chi tiết khuyến mãi
    return redirect('CTKhuyenMai', makm=makm)

def suaCTKhuyenMai(request, makm, masp):
    if request.method == 'POST':
        start_date = request.POST.get('NgayBD')
        end_date = request.POST.get('NgayKT')

        # Kiểm tra xem sản phẩm có tồn tại trong bảng CTKhuyenMai không
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 1 FROM CTKhuyenMai WHERE MaKM = %s AND MaSP = %s
            """, [makm, masp])
            
            if cursor.fetchone():  # Nếu tìm thấy sản phẩm khuyến mãi này
                # Thực hiện câu lệnh UPDATE để cập nhật NgayBD và NgayKT
                cursor.execute("""
                    UPDATE CTKhuyenMai
                    SET NgayBD = %s, NgayKT = %s
                    WHERE MaKM = %s AND MaSP = %s
                """, [start_date, end_date, makm, masp])
                
                messages.success(request, "Cập nhật thông tin khuyến mãi thành công!")
            else:
                messages.error(request, "Không tìm thấy sản phẩm trong khuyến mãi.")

        # Sau khi thực hiện cập nhật, quay lại trang chi tiết khuyến mãi
        return HttpResponseRedirect(reverse('CTKhuyenMai', kwargs={'makm': makm}))
    
    return render(request, 'admin/suaCTKhuyenMai.html', {'makm': makm, 'masp': masp})





# #--------------------------------------------------------QUẢN LÝ NHÀ CUNG CẤP-----------------------------------------------# 
#tạo mã tự động 
def taoMaNCC():
    cursor = connection.cursor()
    query = "SELECT MAX(CAST(SUBSTRING(MaNCC, 4, LENGTH(MaNCC) - 3) AS UNSIGNED)) FROM NhaCungCap"
    cursor.execute(query)
    result = cursor.fetchone()
    cursor.close()

    if result[0]:
        max_number = int(result[0]) + 1
    else:
        max_number = 1 

    # Tạo mã mới với tiền tố "NCC" và số có 3 chữ số (ví dụ: "NCC001")
    return f"NCC{str(max_number).zfill(3)}"
#danh sách nhà cung cấp 
def danhSachNhaCungCap(request):
    # session_info = is_logged_in(request)
    # if not session_info['email'] or session_info['role'] != 'admin':
    #     return HttpResponseRedirect('/login')
    
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM NhaCungCap")
        listUser = cursor.fetchall()

    # PHÂN TRANG
    items_per_page = 5  # Số sản phẩm trên mỗi trang
    paginator = Paginator(listUser, items_per_page)

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    total_pages = paginator.num_pages

    previous_page = max(page_obj.number - 1, 1)  # Trang trước đó
    next_page = min(page_obj.number + 1, total_pages)  # Trang tiếp theo

    start_page = max(1, min(page_obj.number - 1, total_pages - 2))  # Vị trí của nút trang đầu tiên
    end_page = min(total_pages, start_page + 2)

    page_range = range(start_page, end_page + 1)  # Create a range or list of page numbers

    data = {
        'listUser': page_obj,
        'page_range': page_range,  # Pass the page range to the template
        'total_pages': total_pages,
        'previous_page': previous_page,
        'next_page': next_page,
    }
    return render(request, 'admin/danhSachNhaCungCap.html', data)

# #thêm nhà cung cấp
def themNhaCungCap(request):
    if request.method == 'POST':
        # Sinh mã nhà cung cấp tự động
        ma_ncc = taoMaNCC()
        
        # Lấy dữ liệu từ form
        ten_ncc = request.POST.get('TenNCC')
        dia_chi = request.POST.get('DiaChi')
        sdt = request.POST.get('SDT') 
        email = request.POST.get('Email')
        website = request.POST.get('Website')

        # Thực hiện lệnh INSERT
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO NhaCungCap (MaNCC, TenNCC, DiaChi, SDT, Email, Website)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                    [ma_ncc, ten_ncc, dia_chi, sdt, email, website]
                )
            return redirect('/admin/danhSachNhaCungCap') 
        except Exception as e:
            return HttpResponse(f"Lỗi khi thêm nhà cung cấp: {e}")

    return render(request, 'admin/themNhaCungCap.html')


def suaNhaCungCap(request, maNCC):
    if request.method == 'POST':
        ten_nha_cung_cap = request.POST.get('TenNCC')
        dia_chi = request.POST.get('DiaChi')
        sdt = request.POST.get('SDT')
        email = request.POST.get('Email')
        cursor = connection.cursor()
        query = """
            UPDATE NhaCungCap
            SET TenNCC = %s, DiaChi = %s, SDT = %s, Email = %s
            WHERE MaNCC = %s
        """
        cursor.execute(query, [ten_nha_cung_cap, dia_chi, sdt,email,maNCC])
        cursor.close()

        # Chuyển hướng sau khi cập nhật thành công
        return HttpResponseRedirect('/admin/danhSachNhaCungCap')

    # Lấy thông tin khuyến mãi hiện tại để hiển thị trong form
    cursor = connection.cursor()
    query = "SELECT MaNCC, TenNCC, DiaChi,SDT,Email FROM NhaCungCap WHERE MaNCC = %s"
    cursor.execute(query, [maNCC])
    nha_cung_cap = cursor.fetchone()
    cursor.close()

    # Truyền dữ liệu khuyến mãi vào template
    return render(request, 'admin/suaNhaCungCap.html', {'NhaCungCap': nha_cung_cap})

def xoaNhaCungCap(request, mancc):
    cursor = connection.cursor()
    query = "DELETE FROM NhaCungCap WHERE MaNCC = %s"
    cursor.execute(query, [mancc])
    cursor.close()

    # Chuyển hướng sau khi xóa thành công
    return HttpResponseRedirect('/admin/danhSachNhaCungCap')
##---------------------------------QUẢN LÝ LOẠI SẢN PHẨM------------------------------
#tạo mã tự động 
def taoMaLoaiSP():
    cursor = connection.cursor()
    query = "SELECT MAX(CAST(SUBSTRING(MaLoai, 4, LENGTH(MaLoai) - 3) AS UNSIGNED)) FROM LoaiSanPhan"
    cursor.execute(query)
    result = cursor.fetchone()
    cursor.close()

    if result[0]:
        max_number = int(result[0]) + 1
    else:
        max_number = 1 

    # Tạo mã mới với tiền tố "LSP" và số có 3 chữ số (ví dụ: "LSP001")
    return f"LSP{str(max_number).zfill(3)}"
#danh sách nhà cung cấp 
def danhSachLoaiSP(request):
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT loai.MaLoai, loai.TenLoai, ncc.TenNCC
            FROM LoaiSanPhan loai
            JOIN NhaCungCap ncc ON loai.MaNCC = ncc.MaNCC
        """)
        listLoaiSP = cursor.fetchall()

    # PHÂN TRANG
    items_per_page = 5  # Số sản phẩm trên mỗi trang
    paginator = Paginator(listLoaiSP, items_per_page)

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    total_pages = paginator.num_pages

    previous_page = max(page_obj.number - 1, 1)  # Trang trước đó
    next_page = min(page_obj.number + 1, total_pages)  # Trang tiếp theo

    start_page = max(1, min(page_obj.number - 1, total_pages - 2))  # Vị trí của nút trang đầu tiên
    end_page = min(total_pages, start_page + 2)

    page_range = range(start_page, end_page + 1)  # Tạo một range cho các số trang

    data = {
        'listLoaiSP': page_obj,
        'page_range': page_range,  # Pass the page range to the template
        'total_pages': total_pages,
        'previous_page': previous_page,
        'next_page': next_page,
    }
    return render(request, 'admin/danhSachLoaiSP.html', data)
# #thêm loại sản phẩm
def themLoaiSP(request):
    if request.method == 'POST':
        # Sinh mã loại sản phẩm tự động
        ma_lsp = taoMaLoaiSP()
        
        # Lấy dữ liệu từ form
        ten_loai = request.POST.get('TenLoai')
        ma_ncc = request.POST.get('MaNCC')

        # Thực hiện lệnh INSERT
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO LoaiSanPhan (MaLoai, TenLoai, MaNCC)
                    VALUES (%s, %s, %s)
                    """,
                    [ma_lsp, ten_loai, ma_ncc]
                )

            # Sau khi thêm thành công, lấy lại danh sách loại sản phẩm
            with connection.cursor() as cursor:
                cursor.execute("SELECT * FROM LoaiSanPhan")
                listLoaiSP = cursor.fetchall()

            # PHÂN TRANG
            items_per_page = 5  # Số sản phẩm trên mỗi trang
            paginator = Paginator(listLoaiSP, items_per_page)

            page_number = request.GET.get('page', 1)
            page_obj = paginator.get_page(page_number)

            total_pages = paginator.num_pages

            previous_page = max(page_obj.number - 1, 1)  # Trang trước đó
            next_page = min(page_obj.number + 1, total_pages)  # Trang tiếp theo

            start_page = max(1, min(page_obj.number - 1, total_pages - 2))  # Vị trí của nút trang đầu tiên
            end_page = min(total_pages, start_page + 2)

            page_range = range(start_page, end_page + 1)  # Create a range or list of page numbers

            # Trả về trang danh sách loại sản phẩm với dữ liệu đã được cập nhật
            data = {
                'listLoaiSP': page_obj,
                'page_range': page_range,  # Pass the page range to the template
                'total_pages': total_pages,
                'previous_page': previous_page,
                'next_page': next_page,
            }
            return render(request, 'admin/danhSachLoaiSP.html', data)
        except Exception as e:
            return HttpResponse(f"Lỗi khi thêm loại sản phẩm: {e}")

    # Lấy danh sách nhà cung cấp (MaNCC và TenNCC) từ bảng NhaCungCap
    with connection.cursor() as cursor:
        cursor.execute("SELECT MaNCC, TenNCC FROM NhaCungCap")
        nha_cung_caps = cursor.fetchall()

    # Gửi danh sách nhà cung cấp đến template
    return render(request, 'admin/themLoaiSP.html', {'nha_cung_caps': nha_cung_caps})


def suaLoaiSP(request, maloai):
    if request.method == 'POST':
        ten_loai = request.POST.get('TenLoai')
        ma_NCC = request.POST.get('MaNCC')

        cursor = connection.cursor()
        query = """
            UPDATE LoaiSanPhan
            SET TenLoai = %s, MaNCC = %s
            WHERE MaLoai = %s
        """
        cursor.execute(query, [ten_loai, ma_NCC, maloai])  # Thêm maloai ở đây
        cursor.close()

        # Chuyển hướng sau khi cập nhật thành công
        return HttpResponseRedirect('/admin/danhSachLoaiSP')

    # Lấy thông tin loại sản phẩm hiện tại để hiển thị trong form
    cursor = connection.cursor()
    query = "SELECT MaLoai, TenLoai, MaNCC FROM LoaiSanPhan WHERE MaLoai = %s"
    cursor.execute(query, [maloai])
    loai_sp = cursor.fetchone()
    cursor.close()

    # Lấy thông tin nhà cung cấp
    cursor = connection.cursor()
    query_ncc = "SELECT MaNCC, TenNCC FROM NhaCungCap"
    cursor.execute(query_ncc)
    nha_cung_caps = cursor.fetchall()
    cursor.close()

    # Truyền dữ liệu vào template
    return render(request, 'admin/suaLoaiSP.html', {
        'LoaiSanPhan': loai_sp,
        'nha_cung_caps': nha_cung_caps
    })


def xoaLoaiSP(request, malsp):
    cursor = connection.cursor()
    query = "DELETE FROM LoaiSanPhan WHERE MaLoai = %s"
    cursor.execute(query, [malsp])
    cursor.close()

    # Chuyển hướng sau khi xóa thành công
    return HttpResponseRedirect('/admin/danhSachLoaiSP')

##---------------------------------QUẢN LÝ THƯƠNG HIỆU------------------------------
#tạo mã tự động 
def taoMaTH():
    cursor = connection.cursor()
    query = "SELECT MAX(CAST(SUBSTRING(MaTH, 3, LENGTH(MaTH) - 2) AS UNSIGNED)) FROM ThuongHieu"
    cursor.execute(query)
    result = cursor.fetchone()
    cursor.close()

    if result[0]:
        max_number = int(result[0]) + 1
    else:
        max_number = 1 

    # Tạo mã mới với tiền tố "TH" và số có 3 chữ số (ví dụ: "TH001")
    return f"TH{str(max_number).zfill(3)}"
#danh sách thương hiệu
def danhSachTH(request):
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT th.MaTH, th.TenTH, ncc.TenNCC
            FROM ThuongHieu th
            JOIN NhaCungCap ncc ON th.MaNCC = ncc.MaNCC
        """)
        listTH = cursor.fetchall()

    # PHÂN TRANG
    items_per_page = 5  # Số sản phẩm trên mỗi trang
    paginator = Paginator(listTH, items_per_page)

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    total_pages = paginator.num_pages

    previous_page = max(page_obj.number - 1, 1)  # Trang trước đó
    next_page = min(page_obj.number + 1, total_pages)  # Trang tiếp theo

    start_page = max(1, min(page_obj.number - 1, total_pages - 2))  # Vị trí của nút trang đầu tiên
    end_page = min(total_pages, start_page + 2)

    page_range = range(start_page, end_page + 1)  # Tạo một range cho các số trang

    data = {
        'listTH': page_obj,
        'page_range': page_range,  # Pass the page range to the template
        'total_pages': total_pages,
        'previous_page': previous_page,
        'next_page': next_page,
    }
    return render(request, 'admin/danhSachTH.html', data)
# #thêm Thương hiệu
def themTH(request):
    if request.method == 'POST':
        # Sinh mã thương hiệu tự động
        ma_th = taoMaTH()
        
        # Lấy dữ liệu từ form
        ten_th = request.POST.get('TenTH')
        ma_ncc = request.POST.get('MaNCC')

        # Thực hiện lệnh INSERT
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO ThuongHieu (MaTH, TenTH, MaNCC)
                    VALUES (%s, %s, %s)
                    """,
                    [ma_th, ten_th, ma_ncc]
                )

            # Sau khi thêm thành công, lấy lại danh sách thương hiệu
            with connection.cursor() as cursor:
                cursor.execute("SELECT * FROM ThuongHieu")
                listTH = cursor.fetchall()

            # PHÂN TRANG
            items_per_page = 5  # Số sản phẩm trên mỗi trang
            paginator = Paginator(listTH, items_per_page)

            page_number = request.GET.get('page', 1)
            page_obj = paginator.get_page(page_number)

            total_pages = paginator.num_pages

            previous_page = max(page_obj.number - 1, 1)  # Trang trước đó
            next_page = min(page_obj.number + 1, total_pages)  # Trang tiếp theo

            start_page = max(1, min(page_obj.number - 1, total_pages - 2))  # Vị trí của nút trang đầu tiên
            end_page = min(total_pages, start_page + 2)

            page_range = range(start_page, end_page + 1)  # Create a range or list of page numbers

            # Trả về trang danh sách loại sản phẩm với dữ liệu đã được cập nhật
            data = {
                'listTH': page_obj,
                'page_range': page_range,  # Pass the page range to the template
                'total_pages': total_pages,
                'previous_page': previous_page,
                'next_page': next_page,
            }
            return render(request, 'admin/danhSachTH.html', data)
        except Exception as e:
            return HttpResponse(f"Lỗi khi thêm thương hiệu: {e}")

    # Lấy danh sách nhà cung cấp (MaNCC và TenNCC) từ bảng NhaCungCap
    with connection.cursor() as cursor:
        cursor.execute("SELECT MaNCC, TenNCC FROM NhaCungCap")
        nha_cung_caps = cursor.fetchall()

    # Gửi danh sách nhà cung cấp đến template
    return render(request, 'admin/themTH.html', {'nha_cung_caps': nha_cung_caps})


def suaTH(request, mathuonghieu):
    if request.method == 'POST':
        ten_th = request.POST.get('TenTH')
        ma_NCC = request.POST.get('MaNCC')

        cursor = connection.cursor()
        query = """
            UPDATE ThuongHieu
            SET TenTH = %s, MaNCC = %s
            WHERE MaTH = %s
        """
        cursor.execute(query, [ten_th, ma_NCC, mathuonghieu])  
        cursor.close()

        # Chuyển hướng sau khi cập nhật thành công
        return HttpResponseRedirect('/admin/danhSachTH')

    # Lấy thông tin loại sản phẩm hiện tại để hiển thị trong form
    cursor = connection.cursor()
    query = "SELECT MaTH, TenTH, MaNCC FROM ThuongHieu WHERE MaTH = %s"
    cursor.execute(query, [mathuonghieu])
    thuong_hieu = cursor.fetchone()
    cursor.close()

    # Lấy thông tin nhà cung cấp
    cursor = connection.cursor()
    query_ncc = "SELECT MaNCC, TenNCC FROM NhaCungCap"
    cursor.execute(query_ncc)
    nha_cung_caps = cursor.fetchall()
    cursor.close()

    # Truyền dữ liệu vào template
    return render(request, 'admin/suaTH.html', {
        'ThuongHieu': thuong_hieu,
        'nha_cung_caps': nha_cung_caps
    })


def xoaTH(request, math):
    cursor = connection.cursor()
    query = "DELETE FROM ThuongHieu WHERE MaTH = %s"
    cursor.execute(query, [math])
    cursor.close()

    # Chuyển hướng sau khi xóa thành công
    return HttpResponseRedirect('/admin/danhSachTH')


##---------------------------------QUẢN LÝ SẢN PHẨM------------------------------
def taoMaSP(MaLoai):
    cursor = connection.cursor()
    try:
        # Lấy số lớn nhất theo `MaLoai`
        query = """
            SELECT MAX(CAST(SUBSTR(MaSP, -3) AS UNSIGNED)) AS MaxLast3Digits
            FROM SanPham
            WHERE SUBSTR(MaSP, 3, LENGTH(MaSP) - 5) = %s;
        """
        cursor.execute(query, [MaLoai])
        result = cursor.fetchone()

        if result[0]:
            max_number = int(result[0]) + 1
        else:
            max_number = 1

        return f"SP{MaLoai}{str(max_number).zfill(3)}"
    finally:
        cursor.close()



#danh sách sản phẩm
def danhSachSanPham(request):
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT sp.MaSP, sp.TenSP, sp.GiaBan, sp.SoLuongTon, th.TenTH, lsp.TenLoai
            FROM SanPham sp
            JOIN LoaiSanPhan lsp ON sp.MaLoai = lsp.MaLoai
            JOIN ThuongHieu th ON sp.MaTH = th.MaTH
        """)

        listSP = cursor.fetchall()

    # PHÂN TRANG
    items_per_page = 5  # Số sản phẩm trên mỗi trang
    paginator = Paginator(listSP, items_per_page)

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    total_pages = paginator.num_pages

    previous_page = max(page_obj.number - 1, 1)  # Trang trước đó
    next_page = min(page_obj.number + 1, total_pages)  # Trang tiếp theo

    start_page = max(1, min(page_obj.number - 1, total_pages - 2))  # Vị trí của nút trang đầu tiên
    end_page = min(total_pages, start_page + 2)

    page_range = range(start_page, end_page + 1)  # Tạo một range cho các số trang

    data = {
        'listSP': page_obj,
        'page_range': page_range,  # Pass the page range to the template
        'total_pages': total_pages,
        'previous_page': previous_page,
        'next_page': next_page,
    }
    return render(request, 'admin/danhSachSanPham.html', data)


def taoMaHinh():
    cursor = connection.cursor()
    query = "SELECT MAX(CAST(SUBSTRING(MaHinh, 4, LENGTH(MaHinh) - 3) AS UNSIGNED)) FROM HinhAnh"
    cursor.execute(query)
    result = cursor.fetchone()
    cursor.close()

    if result[0]:
        max_number = int(result[0]) + 1
    else:
        max_number = 1  

    return f"PNH{str(max_number).zfill(3)}"

# ##thêm sản phẩm
def themSP(request):
    if request.method == 'POST':
        ma_loai = request.POST.get('MaLoai')
        # Sinh mã sản phẩm tự động
        ma_sp = taoMaSP(ma_loai)

        # Lấy dữ liệu từ form
        ten_sp = request.POST.get('TenSP')
        nsx = request.POST.get('NSX')
        hsd = request.POST.get('HSD')
        mo_ta = request.POST.get('MoTa')
        gia_ban = request.POST.get('GiaBan')
        so_luong_ton = request.POST.get('SoLuongTon')
        trong_luong = request.POST.get('TrongLuong')
        ma_th = request.POST.get('MaTH')
        hinh_anh = request.FILES.get('image')

        # Kiểm tra dữ liệu hợp lệ trước khi thêm
        if not (ten_sp and nsx and hsd and gia_ban and so_luong_ton and trong_luong and ma_loai and ma_th):
            return HttpResponse("Vui lòng nhập đầy đủ thông tin cần thiết")

        # Thực hiện lệnh INSERT
        try:
            # Lưu hình ảnh vào thư mục uploads nếu có
            if hinh_anh:
                fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'static/img'))
                filename = fs.save(hinh_anh.name, hinh_anh)
                image_url = fs.url(filename)
            else:
                image_url = None

            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO SanPham (MaSP, TenSP, NSX, HSD, MoTa, GiaBan, SoLuongTon, TrongLuong, MaLoai, MaTH, HinhAnh)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    [ma_sp, ten_sp, nsx, hsd, mo_ta, gia_ban, so_luong_ton, trong_luong, ma_loai, ma_th, image_url]
                )
                

            # Sau khi thêm thành công, lấy lại danh sách sản phẩm
            with connection.cursor() as cursor:
                cursor.execute("SELECT * FROM SanPham")
                listSP = cursor.fetchall()

            # PHÂN TRANG
            items_per_page = 5  # Số sản phẩm trên mỗi trang
            paginator = Paginator(listSP, items_per_page)

            page_number = request.GET.get('page', 1)
            page_obj = paginator.get_page(page_number)

            total_pages = paginator.num_pages
            previous_page = max(page_obj.number - 1, 1)
            next_page = min(page_obj.number + 1, total_pages)
            start_page = max(1, min(page_obj.number - 1, total_pages - 2))
            end_page = min(total_pages, start_page + 2)
            page_range = range(start_page, end_page + 1)

            data = {
                'listSP': page_obj,
                'page_range': page_range,
                'total_pages': total_pages,
                'previous_page': previous_page,
                'next_page': next_page,
            }
            return render(request, 'admin/danhSachSanPham.html', data)
        except Exception as e:
            return HttpResponse(f"Lỗi khi thêm sản phẩm: {e}")

    # Lấy danh sách loại sản phẩm
    with connection.cursor() as cursor:
        cursor.execute("SELECT MaLoai, TenLoai FROM LoaiSanPhan")
        loai_sps = cursor.fetchall()

    # Lấy danh sách thương hiệu
    with connection.cursor() as cursor:
        cursor.execute("SELECT MaTH, TenTH FROM ThuongHieu")
        thuong_hieus = cursor.fetchall()

    return render(request, 'admin/themSP.html', {'loai_sps': loai_sps, 'thuong_hieus': thuong_hieus})










import os
from django.conf import settings
from django.core.files.storage import default_storage
from django.http import HttpResponseRedirect
from django.db import connection
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.core.paginator import Paginator

def suaSP(request, masanpham):
    if request.method == 'POST':
        # Lấy thông tin từ form
        ten_sp = request.POST.get('TenSP')
        hinh_anh = request.FILES.get('HinhAnh')  
        nsx = request.POST.get('NSX')
        hsd = request.POST.get('HSD')
        mo_ta = request.POST.get('MoTa')
        gia_ban = request.POST.get('GiaBan')
        so_luong_ton = request.POST.get('SoLuongTon')
        trong_luong = request.POST.get('TrongLuong')
        ma_loai = request.POST.get('MaLoai')
        ma_th = request.POST.get('MaTH')

        # Xử lý lưu file hình ảnh nếu có
        hinh_anh_path = None
        if hinh_anh:
            # Lưu hình ảnh vào thư mục "uploads/"
            hinh_anh_path = default_storage.save(os.path.join('', hinh_anh.name), hinh_anh)

        # Tạo câu lệnh SQL cho việc cập nhật
        cursor = connection.cursor()
        if hinh_anh_path:
            # Nếu có hình ảnh, cập nhật trường HinhAnh
            query = """
                UPDATE SanPham
                SET TenSP = %s, HinhAnh = %s, NSX = %s, HSD = %s, MoTa = %s, GiaBan = %s, SoLuongTon = %s, TrongLuong = %s, MaLoai = %s, MaTH = %s
                WHERE MaSP = %s
            """
            cursor.execute(query, [
                ten_sp, 
                hinh_anh_path,  # Đường dẫn file đã lưu
                nsx,
                hsd,
                mo_ta,
                gia_ban,
                so_luong_ton,
                trong_luong,
                ma_loai,
                ma_th,
                masanpham
            ])
        else:
            # Nếu không có hình ảnh, không cập nhật trường HinhAnh
            query = """
                UPDATE SanPham
                SET TenSP = %s, NSX = %s, HSD = %s, MoTa = %s, GiaBan = %s, SoLuongTon = %s, TrongLuong = %s, MaLoai = %s, MaTH = %s
                WHERE MaSP = %s
            """
            cursor.execute(query, [
                ten_sp, 
                nsx,
                hsd,
                mo_ta,
                gia_ban,
                so_luong_ton,
                trong_luong,
                ma_loai,
                ma_th,
                masanpham
            ])
        cursor.close()

        # Chuyển hướng sau khi cập nhật thành công
        return HttpResponseRedirect('/admin/danhSachSanPham')

    # Lấy thông tin sản phẩm hiện tại để hiển thị trong form
    cursor = connection.cursor()
    query = "SELECT MaSP, TenSP, NSX, HSD, MoTa, GiaBan, SoLuongTon, TrongLuong, MaLoai, MaTH FROM SanPham WHERE MaSP = %s"
    cursor.execute(query, [masanpham])
    san_pham = cursor.fetchone()
    cursor.close()

    # Lấy thông tin các loại sản phẩm
    cursor = connection.cursor()
    query_loai = "SELECT MaLoai, TenLoai FROM LoaiSanPhan"
    cursor.execute(query_loai)
    loai_sps = cursor.fetchall()
    cursor.close()

    # Lấy thông tin thương hiệu
    cursor = connection.cursor()
    query_th = "SELECT MaTH, TenTH FROM ThuongHieu"
    cursor.execute(query_th)
    thuong_hieus = cursor.fetchall()
    cursor.close()

    # Truyền dữ liệu vào template
    return render(request, 'admin/suaSP.html', {
        'SanPham': san_pham,
        'loai_sps': loai_sps,
        'thuong_hieus': thuong_hieus
    })


def xoaSP(request, masp):
    cursor = connection.cursor()
    query = "DELETE FROM SanPham WHERE MaSP = %s"
    cursor.execute(query, [masp])
    cursor.close()

    # Chuyển hướng sau khi xóa thành công
    return HttpResponseRedirect('/admin/danhSachSanPham')


def xemChiTietSP(request, MaSP):
    SanPham = None
  
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT sp.MaSP, sp.TenSP, sp.HinhAnh, sp.NSX, sp.HSD, sp.MoTa, sp.GiaBan, sp.SoLuongTon, sp.TrongLuong, th.TenTH, lsp.TenLoai
                FROM SanPham sp
                JOIN LoaiSanPhan lsp ON sp.MaLoai = lsp.MaLoai
                JOIN ThuongHieu th ON sp.MaTH = th.MaTH
                WHERE sp.MaSP = %s
            """, [MaSP])
            SanPham = cursor.fetchone()
    except Exception as e:
        print(f"Error when fetching product details: {e}")
        return HttpResponse("Có lỗi xảy ra khi xem chi tiết sản phẩm.")

    if SanPham is None:
        return HttpResponse("Không tìm thấy sản phẩm.")

    return render(request, 'admin/xemCTSanPham.html', {
        'SanPham': SanPham
    })

# #--------------------------------------------------------QUẢN LÝ NGƯỜI DÙNG-----------------------------------------------# 
#danh sách người dùng
def danhSachTaiKhoan(request):
    # session_info = is_logged_in(request)
    # if not session_info['email'] or session_info['role'] != 'admin':
    #     return HttpResponseRedirect('/login')
    
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM TaiKhoan")
        listUser = cursor.fetchall()

    # PHÂN TRANG
    items_per_page = 5  # Số sản phẩm trên mỗi trang
    paginator = Paginator(listUser, items_per_page)

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    total_pages = paginator.num_pages

    previous_page = max(page_obj.number - 1, 1)  # Trang trước đó
    next_page = min(page_obj.number + 1, total_pages)  # Trang tiếp theo

    start_page = max(1, min(page_obj.number - 1, total_pages - 2))  # Vị trí của nút trang đầu tiên
    end_page = min(total_pages, start_page + 2)

    page_range = range(start_page, end_page + 1)  # Create a range or list of page numbers

    data = {
        'listUser': page_obj,
        'page_range': page_range,  # Pass the page range to the template
        'total_pages': total_pages,
        'previous_page': previous_page,
        'next_page': next_page,
    }
    return render(request, 'admin/danhSachTaiKhoan.html', data)

# #thêm người dùng
def themTaiKhoan(request):
    if request.method == 'POST':
        # Lấy dữ liệu từ form
        tenTaiKhoan = request.POST.get('TenTK')
        matKhau = request.POST.get('MatKhau')
        email = request.POST.get('Email')
        vaiTro = request.POST.get('VaiTro')

        # Mã hóa mật khẩu
        hashed_password = make_password(matKhau)

        # Thực hiện lệnh INSERT
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO TaiKhoan (TenTK, MatKhau, Email, VaiTro)
                    VALUES (%s, %s, %s, %s)
                    """,
                    [tenTaiKhoan, hashed_password, email, vaiTro]
                )
            return HttpResponseRedirect('/admin/danhSachTaiKhoan')
        except Exception as e:
            return HttpResponse('Đã có lỗi xảy ra: ' + str(e))
    return render(request, 'admin/themTaiKhoan.html')

def suaTaiKhoan(request, tenTaiKhoan):
    if request.method == 'POST':
        # Lấy dữ liệu từ form
        matKhau = request.POST.get('MatKhau')
        email = request.POST.get('Email')
        vaiTro = request.POST.get('VaiTro')

        # Mã hóa mật khẩu
        hashed_password = make_password(matKhau)

        # Thực hiện lệnh UPDATE
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    UPDATE TaiKhoan
                    SET MatKhau = %s, Email = %s, VaiTro = %s
                    WHERE TenTK = %s
                    """,
                    [hashed_password, email, vaiTro, tenTaiKhoan]
                )
            return HttpResponseRedirect('/admin/danhSachTaiKhoan')
        except Exception as e:
            return HttpResponse('Đã có lỗi xảy ra: ' + str(e))
    else:
        # Lấy thông tin tài khoản hiện tại
        with connection.cursor() as cursor:
            cursor.execute("SELECT TenTK, MatKhau, Email, VaiTro FROM TaiKhoan WHERE TenTK = %s", [tenTaiKhoan])
            user = cursor.fetchone()
            if user is None:
                return HttpResponse('Không tìm thấy tài khoản')

        data = {
            'TenTK': user[0],
            'MatKhau': user[1],
            'Email': user[2],
            'VaiTro': user[3],
        }
        return render(request, 'admin/suaTaiKhoan.html', data)
    
def xoaTaiKhoan(request, tenTaiKhoan):
    if request.method == 'POST':
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM TaiKhoan WHERE TenTK = %s", [tenTaiKhoan])
        return HttpResponseRedirect('/admin/danhSachTaiKhoan')
    else:
        return HttpResponse('Không thể xóa tài khoản')
    
    


##---------------------------------QUẢN LÝ NHẬP HÀNG ------------------------------
##Trang danh sách nhập hàng
def danhSachNhapHang(request):
    # Lấy dữ liệu từ database và tính tổng tiền cho mỗi phiếu nhập hàng
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT 
                pnh.maPNH, 
                ncc.TenNCC, 
                nv.TenNV, 
                pnh.NgayNhap, 
                (SELECT 
                    SUM(ctnh.SoLuong * ctnh.DonGia) 
                    FROM CTPhieuNhapHang ctnh 
                    WHERE ctnh.MaPNH = pnh.MaPNH) AS TongTien
            FROM PhieuNhapHang pnh
            JOIN NhaCungCap ncc ON pnh.MaNCC = ncc.MaNCC
            JOIN TaiKhoan tk ON pnh.TenTK = tk.TenTK
            JOIN NhanVien nv ON tk.TenTK = nv.TenTK
        """)
        listUser = cursor.fetchall()

    # PHÂN TRANG
    items_per_page = 5  # Số phiếu nhập hàng trên mỗi trang
    paginator = Paginator(listUser, items_per_page)

    page_number = request.GET.get('page', 1)  # Lấy số trang hiện tại từ query parameter
    page_obj = paginator.get_page(page_number)

    total_pages = paginator.num_pages

    previous_page = max(page_obj.number - 1, 1)  # Trang trước đó
    next_page = min(page_obj.number + 1, total_pages)  # Trang tiếp theo

    # Xác định phạm vi hiển thị số trang
    start_page = max(1, min(page_obj.number - 1, total_pages - 2))
    end_page = min(total_pages, start_page + 2)

    page_range = range(start_page, end_page + 1)

    # Truyền dữ liệu vào template
    data = {
        'listUser': page_obj,  # Chứa danh sách phiếu nhập hàng cho trang hiện tại
        'page_range': page_range,  # Phạm vi các trang để hiển thị
        'total_pages': total_pages,
        'previous_page': previous_page,
        'next_page': next_page,
    }
    return render(request, 'admin/danhSachNhapHang.html', data)

##Quản lý nhập hàng
def taoMaPNH():
    cursor = connection.cursor()
    query = "SELECT MAX(CAST(SUBSTRING(MaPNH, 4, LENGTH(MaPNH) - 3) AS UNSIGNED)) FROM PhieuNhapHang"
    cursor.execute(query)
    result = cursor.fetchone()
    cursor.close()

    if result[0]:
        max_number = int(result[0]) + 1
    else:
        max_number = 1  

    return f"PNH{str(max_number).zfill(3)}"

def themNhapHang(request):
    maPNH = taoMaPNH()  # Hàm tự động tạo mã phiếu nhập hàng

    if request.method == 'POST':
        # Lấy dữ liệu từ form
        maNCC = request.POST.get('MaNCC')  
        tenTK = request.POST.get('TenTK')  
        ngayNhap_raw = request.POST.get('NgayNhap')  # Dữ liệu từ form, dạng "2024-11-17T19:30"
        tongTien = 0

        # Lấy danh sách sản phẩm được thêm vào
        danhSachSanPham = []
        for key in request.POST.keys():
            if key.startswith("sanpham_"):
                index = key.split("_")[1]
                sanpham_maSP = request.POST.get(f'sanpham_{index}')
                sanpham_soLuong = int(request.POST.get(f'soluong_{index}', 0))
                sanpham_giaNhap = float(request.POST.get(f'gianhap_{index}', 0.0))
                if sanpham_maSP and sanpham_soLuong > 0:
                    danhSachSanPham.append({
                        'MaSP': sanpham_maSP,
                        'SoLuong': sanpham_soLuong,
                        'GiaNhap': sanpham_giaNhap
                    })
                    tongTien += sanpham_soLuong * sanpham_giaNhap

        try:
            # Chuyển đổi `ngayNhap_raw` sang định dạng "YYYY-MM-DD HH:MM:SS"
            datetime_object = datetime.strptime(ngayNhap_raw, "%Y-%m-%dT%H:%M")
            ngayNhap = datetime_object.strftime("%Y-%m-%d %H:%M:%S")
            with connection.cursor() as cursor:
                # Thêm phiếu nhập hàng vào bảng `PhieuNhapHang`
                cursor.execute(
                    """
                    INSERT INTO PhieuNhapHang (MaPNH, MaNCC, TenTK, NgayNhap, TongTien)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    [maPNH, maNCC, tenTK, ngayNhap, tongTien]
                )
                
                # Thêm sản phẩm vào bảng chi tiết `CTPhieuNhapHang`
                for sp in danhSachSanPham:
                    cursor.execute(
                        """
                        INSERT INTO CTPhieuNhapHang (MaPNH, MaSP, SoLuong, DonGia)
                        VALUES (%s, %s, %s, %s)
                        """,
                        [maPNH, sp['MaSP'], sp['SoLuong'], sp['GiaNhap']]
                    )

            return redirect('/admin/danhSachNhapHang')  # Chuyển hướng đến danh sách phiếu nhập hàng
        except Exception as e:
            return HttpResponse('Đã có lỗi xảy ra: ' + str(e))

    # Lấy danh sách nhà cung cấp, nhân viên và sản phẩm để hiển thị trong form
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT MaNCC, TenNCC FROM NhaCungCap")  # Lấy danh sách nhà cung cấp
            nhaCungCapList = cursor.fetchall()

            cursor.execute("SELECT TenTK, TenNV FROM NhanVien")  # Lấy danh sách tên nhân viên
            taiKhoanList = cursor.fetchall()

            cursor.execute("SELECT MaSP, TenSP FROM SanPham")  # Lấy danh sách sản phẩm
            sanPhamList = cursor.fetchall()

            if request.method == 'POST' and maNCC:
                cursor.execute(
                    """
                    SELECT sp.MaSP, sp.TenSP, ls.MaLoai
                    FROM SanPham sp
                    JOIN LoaiSanPham ls ON sp.MaLoai = ls.MaLoai
                    JOIN NhaCungCap ncc ON ls.MaNCC = ncc.MaNCC
                    WHERE ncc.MaNCC = %s
                    """, [maNCC]
                )
                sanPhamList = cursor.fetchall()
    except Exception as e:
        return HttpResponse('Đã có lỗi xảy ra khi lấy danh sách: ' + str(e))

    return render(request, 'admin/themNhapHang.html', {
        'maPNH': maPNH,
        'nhaCungCapList': nhaCungCapList,
        'taiKhoanList': taiKhoanList,
        'sanPhamList': sanPhamList,
    })

def suaNhapHang(request, maPNH):
    if request.method == 'POST':
        maNCC = request.POST.get('MaNCC')
        tenTK = request.POST.get('TenTK')
        ngayNhap_raw = request.POST.get('NgayNhap')
        tongTien = 0

        # Lấy danh sách sản phẩm được chỉnh sửa
        danhSachSanPham = []
        errors = []  # Danh sách lỗi

        for key in request.POST.keys():
            if key.startswith("sanpham_"):
                index = key.split("_")[1]
                sanpham_maSP = request.POST.get(f'sanpham_{index}')
                sanpham_soLuong = int(request.POST.get(f'soluong_{index}', 0))
                sanpham_giaNhap = float(request.POST.get(f'gianhap_{index}', 0.0))

                # Kiểm tra lỗi
                if sanpham_soLuong < 0:
                    errors.append(f"Số lượng sản phẩm '{sanpham_maSP}' không được nhỏ hơn 0.")
                if sanpham_giaNhap < 0:
                    errors.append(f"Đơn giá sản phẩm '{sanpham_maSP}' không được là số âm.")

                if sanpham_maSP and sanpham_soLuong >= 0 and sanpham_giaNhap >= 0:
                    danhSachSanPham.append({
                        'MaSP': sanpham_maSP,
                        'SoLuong': sanpham_soLuong,
                        'GiaNhap': sanpham_giaNhap
                    })
                    tongTien += sanpham_soLuong * sanpham_giaNhap

        if errors:
            # Hiển thị lại form với các lỗi
            return render(request, 'admin/suaNhapHang.html', {
                'errors': errors,
                'danhSachSanPham': danhSachSanPham,
                'form': {
                    'MaNCC': maNCC,
                    'TenTK': tenTK,
                    'NgayNhap': ngayNhap_raw,
                    'TongTien': tongTien
                }
            })

        try:
            datetime_object = datetime.strptime(ngayNhap_raw, "%Y-%m-%dT%H:%M")
            ngayNhap = datetime_object.strftime("%Y-%m-%d %H:%M:%S")

            with connection.cursor() as cursor:
                # Cập nhật phiếu nhập hàng
                cursor.execute(
                    """
                    UPDATE PhieuNhapHang
                    SET MaNCC = %s, TenTK = %s, NgayNhap = %s, TongTien = %s
                    WHERE MaPNH = %s
                    """,
                    [maNCC, tenTK, ngayNhap, tongTien, maPNH]
                )

                # Xóa các sản phẩm cũ trong bảng CTPhieuNhapHang
                cursor.execute(
                    """
                    DELETE FROM CTPhieuNhapHang WHERE MaPNH = %s
                    """,
                    [maPNH]
                )

                # Thêm lại các sản phẩm mới
                for sp in danhSachSanPham:
                    cursor.execute(
                        """
                        INSERT INTO CTPhieuNhapHang (MaPNH, MaSP, SoLuong, DonGia)
                        VALUES (%s, %s, %s, %s)
                        """,
                        [maPNH, sp['MaSP'], sp['SoLuong'], sp['GiaNhap']]
                    )

            return redirect('/admin/danhSachNhapHang')  # Chuyển hướng đến danh sách phiếu nhập hàng
        except Exception as e:
            return HttpResponse('Đã có lỗi xảy ra: ' + str(e))

    else:
        try:
            with connection.cursor() as cursor:
                # Lấy thông tin phiếu nhập hàng hiện tại
                cursor.execute(
                    """
                    SELECT MaPNH, MaNCC, TenTK, NgayNhap, TongTien
                    FROM PhieuNhapHang
                    WHERE MaPNH = %s
                    """,
                    [maPNH]
                )
                phieu_nhap = cursor.fetchone()

                ngay_nhap_formatted = None
                if phieu_nhap[3]:  # Nếu `NgayNhap` không phải None
                    ngay_nhap_formatted = phieu_nhap[3].strftime("%Y-%m-%dT%H:%M")

                cursor.execute(
                    """
                    SELECT sp.MaSP, sp.TenSP, ctpnh.SoLuong, ctpnh.DonGia
                    FROM CTPhieuNhapHang ctpnh
                    JOIN SanPham sp ON ctpnh.MaSP = sp.MaSP
                    WHERE ctpnh.MaPNH = %s
                    """,
                    [maPNH]
                )
                sanPhamTrongPhieu = cursor.fetchall()

                cursor.execute("SELECT MaNCC, TenNCC FROM NhaCungCap")
                nhaCungCapList = cursor.fetchall()

                cursor.execute("SELECT TenTK, TenNV FROM NhanVien")
                taiKhoanList = cursor.fetchall()

                cursor.execute("SELECT MaSP, TenSP FROM SanPham")
                sanPhamList = cursor.fetchall()

            data = {
                'maPNH': phieu_nhap[0],
                'form': {
                    'MaNCC': phieu_nhap[1],
                    'TenTK': phieu_nhap[2],
                    'NgayNhap': ngay_nhap_formatted,
                    'TongTien': phieu_nhap[4],
                },
                'danhSachSanPham': [
                    {
                        'MaSP': sp[0],
                        'TenSP': sp[1],
                        'SoLuong': sp[2],
                        'GiaNhap': sp[3]
                    }
                    for sp in sanPhamTrongPhieu
                ],
                'sanPhamList': sanPhamList,
                'nhaCungCapList': nhaCungCapList,
                'taiKhoanList': taiKhoanList,
            }
            return render(request, 'admin/suaNhapHang.html', data)
        except Exception as e:
            return HttpResponse('Đã có lỗi xảy ra: ' + str(e))


def xoaNhapHang(request, maPNH):
    if request.method == 'POST':
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM PhieuNhapHang WHERE maPNH = %s", [maPNH])
        return HttpResponseRedirect('/admin/danhSachNhapHang')
    else:
        return HttpResponse('Không thể xóa phiếu nhập hàng')

#Xem chi tiết phiếu nhập
def XemCTNhapHang(request, maPNH):
    # Lấy dữ liệu chi tiết từ database
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT 
                ctpnh.MaPNH,
                sp.TenSP,
                ctpnh.SoLuong,
                ctpnh.DonGia
            FROM CTPhieuNhapHang ctpnh
            JOIN PhieuNhapHang pnh ON ctpnh.MaPNH = pnh.MaPNH
            JOIN SanPham sp ON ctpnh.MaSP = sp.MaSP
            WHERE pnh.MaPNH = %s
        """, [maPNH])
        chiTietNhapHang = cursor.fetchall()

    # PHÂN TRANG
    items_per_page = 5  # Số chi tiết nhập hàng trên mỗi trang
    paginator = Paginator(chiTietNhapHang, items_per_page)

    page_number = request.GET.get('page', 1)  # Lấy số trang hiện tại từ query parameter
    page_obj = paginator.get_page(page_number)

    total_pages = paginator.num_pages

    previous_page = max(page_obj.number - 1, 1)  # Trang trước đó
    next_page = min(page_obj.number + 1, total_pages)  # Trang tiếp theo

    # Xác định phạm vi hiển thị số trang
    start_page = max(1, min(page_obj.number - 1, total_pages - 2))
    end_page = min(total_pages, start_page + 2)

    page_range = range(start_page, end_page + 1)

    # Truyền dữ liệu vào template
    data = {
        'listUser': page_obj,  # Danh sách chi tiết nhập hàng cho trang hiện tại
        'page_range': page_range,  # Phạm vi các trang để hiển thị
        'total_pages': total_pages,
        'previous_page': previous_page,
        'next_page': next_page,
        'maPNH': maPNH,  # Mã phiếu nhập hàng
    }
    return render(request, 'admin/xemCTNhapHang.html', data)


# #--------------------------------------------------------QUẢN LÝ ĐƠN HÀNG-----------------------------------------------#
##Lịch sử đơn hàng
def LichSuDonHang(request):
    # Lấy dữ liệu từ database và lọc các đơn hàng đã thanh toán (Trạng thái = 1)
    with connection.cursor() as cursor:
        cursor.execute("""
    SELECT 
        hd.MaHD,
        hd.NgayLap,
        hd.TenTK,
        SUM(cthd.SoLuong * cthd.DonGia) AS TongTien,
        hd.PhuongThucThanhToan,
        hd.DiaChiGiaoHang,
        hd.TenKhachHang,
        hd.SDT,
        CASE 
            WHEN hd.TrangThai = 1 THEN 'Đã thanh toán'
        END AS TrangThai
        FROM HoaDon hd
        LEFT JOIN CTHoaDon cthd ON hd.MaHD = cthd.MaHD
        WHERE hd.TrangThai = 1
        GROUP BY 
            hd.MaHD,
            hd.NgayLap,
            hd.PhuongThucThanhToan,
            hd.DiaChiGiaoHang,
            hd.TenKhachHang,
            hd.TenTK,
            hd.SDT,
            hd.TrangThai
        """)
        listLSHD = cursor.fetchall()

    # PHÂN TRANG
    items_per_page = 5  # Số đơn hàng trên mỗi trang
    paginator = Paginator(listLSHD, items_per_page)

    page_number = request.GET.get('page', 1)  # Lấy số trang hiện tại từ query parameter
    page_obj = paginator.get_page(page_number)

    total_pages = paginator.num_pages

    previous_page = max(page_obj.number - 1, 1)  # Trang trước đó
    next_page = min(page_obj.number + 1, total_pages)  # Trang tiếp theo

    # Xác định phạm vi hiển thị số trang
    start_page = max(1, min(page_obj.number - 1, total_pages - 2))
    end_page = min(total_pages, start_page + 2)

    page_range = range(start_page, end_page + 1)

    # Truyền dữ liệu vào template
    data = {
        'listLSHD': page_obj,  # Chứa danh sách đơn hàng cho trang hiện tại
        'page_range': page_range,  # Phạm vi các trang để hiển thị
        'total_pages': total_pages,
        'previous_page': previous_page,
        'next_page': next_page,
    }
    return render(request, 'admin/LichSuDonHang.html', data)

def XemCTDonHang(request, MaHD):
    # Lấy dữ liệu chi tiết từ database
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT 
                hd.MaHD,
                sp.TenSP,
                cthd.SoLuong,
                cthd.DonGia
            FROM CTHoaDon cthd
            JOIN HoaDon hd ON cthd.MaHD = hd.MaHD
            JOIN SanPham sp ON cthd.MaSP = sp.MaSP
            WHERE hd.MaHD = %s
        """, [MaHD])
        listCTHD = cursor.fetchall()

    # Truyền dữ liệu vào template
    data = {
        'listCTHD': listCTHD,
        'MaHD': MaHD,  # Truyền thêm mã hóa đơn để hiển thị nếu cần
    }
    return render(request, 'admin/XemCTDonHang.html', data)





# #--------------------------------------------------------THỐNG KÊ BÁO CÁO-----------------------------------------------# 
#Chức năng thống kê báo cáo

#Chức năng thống kê báo cáo doanh thu
def thongKeDoanhThu(request):
    with connection.cursor() as cursor:
        # Tổng doanh thu theo ngày
        cursor.execute("""
            SELECT strftime('%Y-%m-%d', NgayLap) AS Ngay, SUM(TongTien) AS TongDoanhThu
            FROM HoaDon
            WHERE TrangThaiThanhToan = 'Đã thanh toán'
            GROUP BY strftime('%Y-%m-%d', NgayLap)
            ORDER BY Ngay;
        """)
        doanh_thu_ngay = cursor.fetchall()

        # Tổng doanh thu theo tháng
        cursor.execute("""
            SELECT strftime('%Y', NgayLap) AS Nam, strftime('%m', NgayLap) AS Thang, SUM(TongTien) AS TongDoanhThu
            FROM HoaDon
            WHERE TrangThaiThanhToan = 'Đã thanh toán'
            GROUP BY strftime('%Y', NgayLap), strftime('%m', NgayLap)
            ORDER BY Nam, Thang;
        """)
        doanh_thu_thang = cursor.fetchall()

        # Tổng doanh thu theo năm
        cursor.execute("""
            SELECT strftime('%Y', NgayLap) AS Nam, SUM(TongTien) AS TongDoanhThu
            FROM HoaDon
            WHERE TrangThaiThanhToan = 'Đã thanh toán'
            GROUP BY strftime('%Y', NgayLap)
            ORDER BY Nam;
        """)
        doanh_thu_nam = cursor.fetchall()

        # Doanh thu từ các loại sản phẩm
        cursor.execute("""
            SELECT lsp.TenLoai, SUM(cthd.SoLuong * cthd.DonGia) AS DoanhThuLoaiSP
            FROM CTHoaDon cthd
            JOIN SanPham sp ON cthd.MaSP = sp.MaSP
            JOIN LoaiSanPhan lsp ON sp.MaLoai = lsp.MaLoai
            JOIN HoaDon hd ON cthd.MaHD = hd.MaHD
            WHERE hd.TrangThaiThanhToan = 'Đã thanh toán'
            GROUP BY lsp.TenLoai
            ORDER BY DoanhThuLoaiSP DESC;
        """)
        doanh_thu_loai_sp = cursor.fetchall()

        # Doanh thu từ các thương hiệu
        cursor.execute("""
            SELECT th.TenTH, SUM(cthd.SoLuong * cthd.DonGia) AS DoanhThuTH
            FROM CTHoaDon cthd
            JOIN SanPham sp ON cthd.MaSP = sp.MaSP
            JOIN ThuongHieu th ON sp.MaTH = th.MaTH
            JOIN HoaDon hd ON cthd.MaHD = hd.MaHD
            WHERE hd.TrangThaiThanhToan = 'Đã thanh toán'
            GROUP BY th.TenTH
            ORDER BY DoanhThuTH DESC;
        """)
        doanh_thu_thuong_hieu = cursor.fetchall()

    context = {
        'doanh_thu_ngay': doanh_thu_ngay,
        'doanh_thu_thang': doanh_thu_thang,
        'doanh_thu_nam': doanh_thu_nam,
        'doanh_thu_loai_sp': doanh_thu_loai_sp,
        'doanh_thu_thuong_hieu': doanh_thu_thuong_hieu,
    }
    return render(request, 'admin/thongKeDoanhThu.html', context)

#Chức năng thống kê báo cáo sản phẩm
def thongKeSanPham(request):
    with connection.cursor() as cursor:
        # Sản phẩm đã bán
        cursor.execute("""
            SELECT sp.TenSP, SUM(cthd.SoLuong) AS SoLuongBan
            FROM CTHoaDon cthd
            JOIN SanPham sp ON cthd.MaSP = sp.MaSP
            JOIN HoaDon hd ON cthd.MaHD = hd.MaHD
            WHERE hd.TrangThaiThanhToan = 'Đã thanh toán'
            GROUP BY sp.TenSP
            ORDER BY SoLuongBan DESC;
        """)
        sp_da_ban = cursor.fetchall()
        sp_ban_chay = sp_da_ban[:10]

        # Sản phẩm tồn kho
        cursor.execute("""
            SELECT sp.TenSP, sp.SoLuongTon
            FROM SanPham sp
            WHERE sp.SoLuongTon > 0
            ORDER BY sp.SoLuongTon DESC;
        """)
        sp_ton_kho = cursor.fetchall()

    context = {
        'sp_da_ban': sp_da_ban,
        'sp_ban_chay': sp_ban_chay,
        'sp_ton_kho': sp_ton_kho,
    }
    return render(request, 'admin/thongKeSanPham.html', context)

#Chức năng thống kê báo cáo đơn hàng
def thongKeDonHang(request):
    with connection.cursor() as cursor:
        # Số lượng đơn hàng theo trạng thái
        cursor.execute("""
            SELECT TrangThai, COUNT(*) AS SoLuongDonHang
            FROM HoaDon
            GROUP BY TrangThai;
        """)
        don_hang_trang_thai = cursor.fetchall()

        # Tổng số đơn hàng
        cursor.execute("""
            SELECT COUNT(*) AS TongDonHang
            FROM HoaDon;
        """)
        tong_so_don_hang = cursor.fetchone()[0] or 0

        # Tỉ lệ đơn hàng chưa được xác nhận
        cursor.execute("""
            SELECT ROUND(SUM(CASE WHEN TrangThai = '0' THEN 1 ELSE 0 END) * 100.0 / COUNT(*), 2) AS TyLeHuyDon
            FROM HoaDon;
        """)
        ty_le_huy_don = cursor.fetchone()[0] or 0

        # Đơn hàng đang xử lý
        cursor.execute("""
            SELECT COUNT(*) AS DangXuLy
            FROM HoaDon
            WHERE TrangThai = '1';
        """)
        don_hang_dang_xu_ly = cursor.fetchone()[0] or 0

        # Đơn hàng theo khách hàng
        cursor.execute("""
            SELECT TenKhachHang, COUNT(MaHD) AS SoDonHang
            FROM HoaDon
            GROUP BY TenKhachHang
            ORDER BY SoDonHang DESC
            LIMIT 5;
        """)
        don_hang_theo_khach_hang = cursor.fetchall()

    context = {
        'don_hang_trang_thai': don_hang_trang_thai,
        'tong_so_don_hang': tong_so_don_hang,
        'ty_le_huy_don': ty_le_huy_don,
        'don_hang_dang_xu_ly': don_hang_dang_xu_ly,
        'don_hang_theo_khach_hang': don_hang_theo_khach_hang,
    }
    return render(request, 'admin/thongKeDonHang.html', context)

# Xuất báo cáo doanh thu
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def export_doanh_thu_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Thống kê Doanh thu"

    # Tiêu đề lớn
    ws.merge_cells('A1:E1')
    ws['A1'] = "Thống kê Doanh thu"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].fill = PatternFill("solid", fgColor="DDDDDD")

    def add_section(title, headers, data, start_row):
        # Tiêu đề phần
        ws.merge_cells(start_row + ':' + start_row.replace('A', 'E'))
        ws[start_row] = title
        ws[start_row].font = Font(bold=True, size=12, color="FFFFFF")
        ws[start_row].fill = PatternFill("solid", fgColor="4F81BD")
        ws[start_row].alignment = Alignment(horizontal="left")

        # Header
        header_row = int(start_row[1:]) + 1
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="B8CCE4")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # Data
        data_row = header_row + 1
        for row in data:
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=data_row, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="right" if isinstance(value, (int, float)) else "left")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
            data_row += 1
        return data_row + 1

    with connection.cursor() as cursor:
        # Doanh thu theo ngày
        cursor.execute("""
            SELECT strftime('%Y-%m-%d', NgayLap) AS Ngay, SUM(TongTien) AS TongDoanhThu
            FROM HoaDon
            WHERE TrangThaiThanhToan = 'Đã thanh toán'
            GROUP BY strftime('%Y-%m-%d', NgayLap)
            ORDER BY Ngay;
        """)
        doanh_thu_ngay = cursor.fetchall()
        next_row = add_section("Doanh thu theo ngày", ["Ngày", "Doanh thu"], doanh_thu_ngay, "A3")

        # Doanh thu theo tháng
        cursor.execute("""
            SELECT strftime('%Y', NgayLap) AS Nam, strftime('%m', NgayLap) AS Thang, SUM(TongTien) AS TongDoanhThu
            FROM HoaDon
            WHERE TrangThaiThanhToan = 'Đã thanh toán'
            GROUP BY strftime('%Y', NgayLap), strftime('%m', NgayLap)
            ORDER BY Nam, Thang;
        """)
        doanh_thu_thang = cursor.fetchall()
        next_row = add_section("Doanh thu theo tháng", ["Năm", "Tháng", "Doanh thu"], doanh_thu_thang, f"A{next_row}")

        # Doanh thu theo năm
        cursor.execute("""
            SELECT strftime('%Y', NgayLap) AS Nam, SUM(TongTien) AS TongDoanhThu
            FROM HoaDon
            WHERE TrangThaiThanhToan = 'Đã thanh toán'
            GROUP BY strftime('%Y', NgayLap)
            ORDER BY Nam;
        """)
        doanh_thu_nam = cursor.fetchall()
        next_row = add_section("Doanh thu theo năm", ["Năm", "Doanh thu"], doanh_thu_nam, f"A{next_row}")

        # Doanh thu theo loại sản phẩm
        cursor.execute("""
            SELECT lsp.TenLoai, SUM(cthd.SoLuong * cthd.DonGia) AS DoanhThuLoaiSP
            FROM CTHoaDon cthd
            JOIN SanPham sp ON cthd.MaSP = sp.MaSP
            JOIN LoaiSanPhan lsp ON sp.MaLoai = lsp.MaLoai
            JOIN HoaDon hd ON cthd.MaHD = hd.MaHD
            WHERE hd.TrangThaiThanhToan = 'Đã thanh toán'
            GROUP BY lsp.TenLoai
            ORDER BY DoanhThuLoaiSP DESC;
        """)
        doanh_thu_loai_sp = cursor.fetchall()
        next_row = add_section("Doanh thu theo loại sản phẩm", ["Loại sản phẩm", "Doanh thu"], doanh_thu_loai_sp, f"A{next_row}")

        # Doanh thu theo thương hiệu
        cursor.execute("""
            SELECT th.TenTH, SUM(cthd.SoLuong * cthd.DonGia) AS DoanhThuTH
            FROM CTHoaDon cthd
            JOIN SanPham sp ON cthd.MaSP = sp.MaSP
            JOIN ThuongHieu th ON sp.MaTH = th.MaTH
            JOIN HoaDon hd ON cthd.MaHD = hd.MaHD
            WHERE hd.TrangThaiThanhToan = 'Đã thanh toán'
            GROUP BY th.TenTH
            ORDER BY DoanhThuTH DESC;
        """)
        doanh_thu_th = cursor.fetchall()
        add_section("Doanh thu theo thương hiệu", ["Thương hiệu", "Doanh thu"], doanh_thu_th, f"A{next_row}")

    # Tạo HTTP Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=thong_ke_doanh_thu.xlsx"
    wb.save(response)
    return response


# Xuất báo cáo sản phẩm
def export_san_pham_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Thống kê Sản phẩm"

    # Tiêu đề chính
    ws.merge_cells('A1:D1')
    ws['A1'] = "Thống kê Sản phẩm"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].fill = PatternFill("solid", fgColor="DDDDDD")

    def add_section(title, headers, data, start_row):
        # Tiêu đề phần
        ws.merge_cells(start_row + ':' + start_row.replace('A', 'D'))
        ws[start_row] = title
        ws[start_row].font = Font(bold=True, size=12, color="FFFFFF")
        ws[start_row].fill = PatternFill("solid", fgColor="4F81BD")
        ws[start_row].alignment = Alignment(horizontal="left")

        # Header
        header_row = int(start_row[1:]) + 1
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="B8CCE4")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # Data
        data_row = header_row + 1
        for row in data:
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=data_row, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="right" if isinstance(value, (int, float)) else "left")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
            data_row += 1
        return data_row + 1

    with connection.cursor() as cursor:
        # Sản phẩm đã bán
        cursor.execute("""
            SELECT sp.TenSP, SUM(cthd.SoLuong) AS SoLuongBan
            FROM CTHoaDon cthd
            JOIN SanPham sp ON cthd.MaSP = sp.MaSP
            JOIN HoaDon hd ON cthd.MaHD = hd.MaHD
            WHERE hd.TrangThaiThanhToan = 'Đã thanh toán'
            GROUP BY sp.TenSP
            ORDER BY SoLuongBan DESC;
        """)
        san_pham_da_ban = cursor.fetchall()
        next_row = add_section(
            "Sản phẩm đã bán",
            ["Tên sản phẩm", "Số lượng bán"],
            san_pham_da_ban,
            "A3"
        )

        # Sản phẩm tồn kho
        cursor.execute("""
            SELECT sp.TenSP, sp.SoLuongTon
            FROM SanPham sp
            WHERE sp.SoLuongTon > 0
            ORDER BY sp.SoLuongTon DESC;
        """)
        san_pham_ton_kho = cursor.fetchall()
        add_section(
            "Sản phẩm tồn kho",
            ["Tên sản phẩm", "Số lượng tồn"],
            san_pham_ton_kho,
            f"A{next_row}"
        )

    # Tạo HTTP Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=thong_ke_san_pham.xlsx"
    wb.save(response)
    return response


# Xuất báo cáo đơn hàng
def export_don_hang_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Thống kê Đơn hàng"

    # Tiêu đề chính
    ws.merge_cells('A1:D1')
    ws['A1'] = "Thống kê Đơn hàng"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].fill = PatternFill("solid", fgColor="DDDDDD")

    def add_section(title, headers, data, start_row):
        # Tiêu đề phần
        ws.merge_cells(start_row + ':' + start_row.replace('A', 'D'))
        ws[start_row] = title
        ws[start_row].font = Font(bold=True, size=12, color="FFFFFF")
        ws[start_row].fill = PatternFill("solid", fgColor="4F81BD")
        ws[start_row].alignment = Alignment(horizontal="left")

        # Header
        header_row = int(start_row[1:]) + 1
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="B8CCE4")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # Data
        data_row = header_row + 1
        for row in data:
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=data_row, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="right" if isinstance(value, (int, float)) else "left")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
            data_row += 1
        return data_row + 1

    with connection.cursor() as cursor:
        # Số lượng đơn hàng theo trạng thái
        cursor.execute("""
            SELECT TrangThai, COUNT(*) AS SoLuongDonHang
            FROM HoaDon
            GROUP BY TrangThai;
        """)
        so_luong_trang_thai = cursor.fetchall()
        next_row = add_section(
            "Số lượng đơn hàng theo trạng thái",
            ["Trạng thái", "Số lượng"],
            so_luong_trang_thai,
            "A3"
        )

        # Tổng số đơn hàng
        cursor.execute("""
            SELECT COUNT(*) AS TongDonHang
            FROM HoaDon;
        """)
        tong_so_don_hang = cursor.fetchone()[0] or 0
        next_row = add_section(
            "Tổng số đơn hàng",
            ["Thông tin", "Giá trị"],
            [["Tổng số đơn hàng", tong_so_don_hang]],
            f"A{next_row}"
        )

        # Tỉ lệ đơn hàng chưa xác nhận
        cursor.execute("""
            SELECT ROUND(SUM(CASE WHEN TrangThai = '0' THEN 1 ELSE 0 END) * 100.0 / COUNT(*), 2) AS TyLeHuyDon
            FROM HoaDon;
        """)
        ty_le_huy_don = cursor.fetchone()[0] or 0
        next_row = add_section(
            "Tỉ lệ đơn hàng chưa xác nhận",
            ["Thông tin", "Tỉ lệ"],
            [["Tỉ lệ đơn hàng chưa xác nhận", f"{ty_le_huy_don}%"]],
            f"A{next_row}"
        )

        # Top 5 khách hàng đặt hàng nhiều nhất
        cursor.execute("""
            SELECT TenKhachHang, COUNT(MaHD) AS SoDonHang
            FROM HoaDon
            GROUP BY TenKhachHang
            ORDER BY SoDonHang DESC
            LIMIT 5;
        """)
        top_khach_hang = cursor.fetchall()
        add_section(
            "Top 5 khách hàng đặt hàng nhiều nhất",
            ["Tên khách hàng", "Số đơn hàng"],
            top_khach_hang,
            f"A{next_row}"
        )

    # Tạo HTTP Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=thong_ke_don_hang.xlsx"
    wb.save(response)
    return response

